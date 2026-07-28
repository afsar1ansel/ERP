from db import db
from sqlalchemy.sql import text
from datetime import datetime
import cloudinary.uploader
import os
from werkzeug.utils import secure_filename
from sqlalchemy import text
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from flask import send_file
from io import BytesIO
from flask import send_file, make_response
import json
from openpyxl.worksheet.datavalidation import DataValidation
from helper.files import send_workbook_response

class VendorClass:
    
    def upload_vendor_logo(self, file):
        """Upload vendor logo to Cloudinary with validation"""
        
        # Validation: File size (2MB limit)
        if len(file.read()) > 2 * 1024 * 1024:  # 2MB in bytes
            file.seek(0)  # Reset file pointer
            return {"errFlag": 1, "message": "File size must be less than 2MB"}
        
        file.seek(0)  # Reset file pointer after checking size
        
        # Validation: File type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
        filename = secure_filename(file.filename)
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return {"errFlag": 1, "message": "Invalid file type. Allowed: PNG, JPG, JPEG, GIF, BMP, WEBP"}
        
        try:
            # Upload to Cloudinary
            upload_result = cloudinary.uploader.upload(
                file,
                folder="vendors",
                transformation=[
                    {'width': 300, 'height': 300, 'crop': 'limit'},
                    {'quality': 'auto'},
                    {'format': 'auto'}
                ]
            )
            
            return {"errFlag": 0, "url": upload_result['secure_url'], "public_id": upload_result['public_id']}
            
        except Exception as e:
            return {"errFlag": 1, "message": "Upload failed"}

    def delete_vendor_logo(self, public_id):
        """Delete logo from Cloudinary"""
        try:
            result = cloudinary.uploader.destroy(public_id)
            if result.get('result') == 'ok':
                return True
            return False
        except:
            return False

    


    def getVendors(self):
        sql = text('''
            SELECT v.* 
            FROM vendors v
            ORDER BY v.vendor_name
        ''')
        with db.engine.connect() as conn:
            vendors = conn.execute(sql).mappings().all()

        if not vendors:
            return []

        vendor_ids = [v['id'] for v in vendors]

        # build placeholders
        placeholders = ",".join([f":id{i}" for i in range(len(vendor_ids))])
        materials_sql = text(f'''
            SELECT
                vrm.vendor_id,
                vrm.raw_material_id,
                rm.material_name,
                rm.material_code,
                rm.material_description,
                rm.raw_material_image,
                rm.raw_material_image_public_id,
                rm.unit_of_measure,
                rm.min_stock_level,
                rm.max_stock_level,
                rm.stock_qty,
                rm.storage_location_id,
                c.category_name AS raw_material_category_name,
                c.id AS raw_material_category_id
            FROM vendor_raw_materials vrm
            JOIN raw_materials rm ON vrm.raw_material_id = rm.id
            LEFT JOIN raw_material_categories c ON rm.raw_material_category_id = c.id
            WHERE vrm.vendor_id IN ({placeholders})
            ORDER BY vrm.vendor_id, rm.material_name
        ''')

        params = {f"id{i}": vendor_ids[i] for i in range(len(vendor_ids))}
        items_by_vendor = defaultdict(list)
        with db.engine.connect() as conn:
            all_items = conn.execute(materials_sql, params).mappings().all()
            for row in all_items:
                items_by_vendor[row['vendor_id']].append(dict(row))

        result = []
        for v in vendors:
            vd = dict(v)
            vd['raw_materials'] = items_by_vendor.get(vd['id'], [])
            result.append(vd)

        return result



    def addVendor(self, vendorName, contactPerson, email, phone,
                 address, city, state, pincode, gstNo, panNo, bankName, accountNo,
                 ifscCode, paymentTerms, creditLimit, notes, adminUserId, vendorLogoFile=None, raw_materials=None):
        """Adds a new vendor with optional vendor logo and raw materials."""
        # Check for duplicate vendor name
        duplicate_check = self.chkDuplicateVendorName(vendorName)
        if duplicate_check:
            return {"errFlag": 1, "message": "A vendor with this name already exists"}

        # Handle vendor logo upload
        vendor_logo_url = ""
        vendor_logo_public_id = ""
        
        if vendorLogoFile and vendorLogoFile.filename != '':
            upload_result = self.upload_vendor_logo(vendorLogoFile)
            if upload_result["errFlag"] == 1:
                return upload_result
            vendor_logo_url = upload_result["url"]
            vendor_logo_public_id = upload_result["public_id"]

        data = {
            'vendorName': vendorName,
            'contactPerson': contactPerson,
            'email': email,
            'phone': phone,
            'address': address,
            'city': city,
            'state': state,
            'pincode': pincode,
            'gstNo': gstNo,
            'panNo': panNo,
            'bankName': bankName,
            'accountNo': accountNo,
            'ifscCode': ifscCode,
            'paymentTerms': paymentTerms,
            'creditLimit': creditLimit,
            'notes': notes,
            'vendorLogo': vendor_logo_url,
            'vendorLogoPublicId': vendor_logo_public_id,
            'status': 1,
            'createdAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'createdAdminId': adminUserId
        }

        try:
            with db.engine.connect() as conn:
                # Start transaction
                trans = conn.begin()
                
                # Insert vendor
                vendor_sql = text('''
                    INSERT INTO vendors (
                        vendor_name, contact_person, email, phone, 
                        address, city, state, pincode, gst_no, pan_no, bank_name, account_no,
                        ifsc_code, payment_terms, credit_limit, notes, vender_logo, vender_logo_public_id,
                        status, created_at, created_admin_id
                    ) 
                    VALUES (
                        :vendorName, :contactPerson, :email, :phone, 
                        :address, :city, :state, :pincode, :gstNo, :panNo, :bankName, :accountNo,
                        :ifscCode, :paymentTerms, :creditLimit, :notes, :vendorLogo, :vendorLogoPublicId,
                        :status, :createdAt, :createdAdminId
                    )
                ''')
                result = conn.execute(vendor_sql, data)
                vendor_id = result.lastrowid
                
                # Add raw materials if provided
                if raw_materials:
                    for material in raw_materials:
                        material_data = {
                            'vendor_id': vendor_id,
                            'raw_material_id': material['raw_material_id']
                        }
                        
                        material_sql = text('''
                            INSERT INTO vendor_raw_materials 
                            (vendor_id, raw_material_id)
                            VALUES (:vendor_id, :raw_material_id)
                        ''')
                        conn.execute(material_sql, material_data)
                
                # Commit transaction
                trans.commit()
                
                return vendor_id
                
        except Exception as e:
            # Rollback on error
            if 'trans' in locals():
                trans.rollback()
            return {"errFlag": 1, "message": "Error while adding vendor"}
    
    def updateVendor(self, vendorId, vendorName, contactPerson, email, phone,
                    address, city, state, pincode, gstNo, panNo, bankName, accountNo,
                    ifscCode, paymentTerms, creditLimit, notes, adminUserId, vendorLogoFile=None, raw_materials=None):
        """Updates an existing vendor and its raw materials."""
        # Check for duplicate vendor name (excluding the current vendor)
        duplicate_check = self.chkDuplicateVendorName(vendorName, vendorId)
        if duplicate_check:
            return {"errFlag": 1, "message": "Another vendor with this name already exists"}

        # Get current vendor details for logo cleanup
        current_vendor = self.getVendorDetails(vendorId)
        current_logo_public_id = current_vendor.get('vender_logo_public_id') if current_vendor else None

        # Handle vendor logo upload
        vendor_logo_url = current_vendor.get('vender_logo') if current_vendor else ""
        vendor_logo_public_id = current_logo_public_id
        
        if vendorLogoFile and vendorLogoFile.filename != '':
            upload_result = self.upload_vendor_logo(vendorLogoFile)
            if upload_result["errFlag"] == 1:
                return upload_result
            vendor_logo_url = upload_result["url"]
            vendor_logo_public_id = upload_result["public_id"]
            
            # Delete old logo if it exists
            if current_logo_public_id:
                self.delete_vendor_logo(current_logo_public_id)

        data = {
            'vendorId': vendorId,
            'vendorName': vendorName,
            'contactPerson': contactPerson,
            'email': email,
            'phone': phone,
            'address': address,
            'city': city,
            'state': state,
            'pincode': pincode,
            'gstNo': gstNo,
            'panNo': panNo,
            'bankName': bankName,
            'accountNo': accountNo,
            'ifscCode': ifscCode,
            'paymentTerms': paymentTerms,
            'creditLimit': creditLimit,
            'notes': notes,
            'vendorLogo': vendor_logo_url,
            'vendorLogoPublicId': vendor_logo_public_id,
            'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'updatedAdminId': adminUserId
        }

        try:
            with db.engine.connect() as conn:
                # Start transaction
                trans = conn.begin()
                
                # Update vendor
                vendor_sql = text('''
                    UPDATE vendors 
                    SET vendor_name = :vendorName,
                        contact_person = :contactPerson,
                        email = :email,
                        phone = :phone,
                        address = :address,
                        city = :city,
                        state = :state,
                        pincode = :pincode,
                        gst_no = :gstNo,
                        pan_no = :panNo,
                        bank_name = :bankName,
                        account_no = :accountNo,
                        ifsc_code = :ifscCode,
                        payment_terms = :paymentTerms,
                        credit_limit = :creditLimit,
                        notes = :notes,
                        vender_logo = :vendorLogo,
                        vender_logo_public_id = :vendorLogoPublicId,
                        updated_at = :updatedAt,
                        updated_admin_id = :updatedAdminId
                    WHERE id = :vendorId
                ''')
                conn.execute(vendor_sql, data)
                
                # Update raw materials if provided
                if raw_materials is not None:
                    # First delete existing materials
                    delete_sql = text('DELETE FROM vendor_raw_materials WHERE vendor_id = :vendor_id')
                    conn.execute(delete_sql, {'vendor_id': vendorId})
                    
                    # Add new materials
                    for material in raw_materials:
                        material_data = {
                            'vendor_id': vendorId,
                            'raw_material_id': material['raw_material_id']
                        }
                        
                        material_sql = text('''
                            INSERT INTO vendor_raw_materials 
                            (vendor_id, raw_material_id)
                            VALUES (:vendor_id, :raw_material_id)
                        ''')
                        conn.execute(material_sql, material_data)
                
                # Commit transaction
                trans.commit()
                
                return 1  # Success
                
        except Exception as e:
            # Rollback on error
            if 'trans' in locals():
                trans.rollback()
            return {"errFlag": 1, "message": "Error while updating vendor"}
    
    def getVendorDetails(self, vendorId):
        """Fetches details for a single vendor including raw materials."""
        try:
            with db.engine.connect() as conn:
                # Get vendor details
                vendor_sql = text('SELECT * FROM vendors WHERE id = :vendorId')
                vendor_data = {'vendorId': vendorId}
                vendor_result = conn.execute(vendor_sql, vendor_data)
                vendor_details = vendor_result.mappings().all()
                
                if not vendor_details:
                    return None
                
                # Get vendor's raw materials
                materials_sql = text('''
                    SELECT vrm.raw_material_id, rm.material_name , c.category_name as raw_material_category_name
                    FROM vendor_raw_materials vrm
                    JOIN raw_materials rm ON vrm.raw_material_id = rm.id
                    JOIN raw_material_categories c ON rm.raw_material_category_id = c.id
                    WHERE vrm.vendor_id = :vendorId
                ''')
                materials_result = conn.execute(materials_sql, vendor_data)
                materials = materials_result.mappings().all()
                
                # Combine results
                vendor_info = dict(vendor_details[0])
                vendor_info['raw_materials'] = [dict(material) for material in materials]
                
                return vendor_info
                
        except Exception as e:
            print("Error getting vendor details:", e)
            return None
    
    def changeVendorStatus(self, vendorId, status, adminUserId):
        """Changes the active/inactive status of a vendor."""
        data = {
            'vendorId': vendorId,
            'status': status,
            'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'updatedAdminId': adminUserId
        }

        sql = text('''
            UPDATE vendors 
            SET status = :status, 
                updated_at = :updatedAt,
                updated_admin_id = :updatedAdminId
            WHERE id = :vendorId
        ''')
        with db.engine.connect() as conn:
            responseData = conn.execute(sql, data)
            conn.commit()
        return responseData.rowcount
    
    def chkDuplicateVendorName(self, vendorName, vendorId=None):
        """Checks for duplicate vendor names to ensure uniqueness."""
        if vendorId:
            sql = text('''
                SELECT * FROM vendors 
                WHERE LOWER(vendor_name) = LOWER(:vendorName) AND id != :vendorId  
            ''')
            data = {'vendorName': vendorName, 'vendorId': vendorId}
        else:
            sql = text('''
                SELECT * FROM vendors 
                WHERE LOWER(vendor_name) = LOWER(:vendorName)
            ''')
            data = {'vendorName': vendorName}
            
        with db.engine.connect() as conn:
            responseData = conn.execute(sql, data)
        
        return responseData.mappings().all()
      ###-----------------Bulk vendor upload-------------- ###
    
    def bulkUploadVendors(self, excel_file, admin_user_id):
        try:
            df = pd.read_excel(excel_file, sheet_name="Vendors")
            
            # Remove rows where all values are NaN (completely empty rows)
            df = df.dropna(how='all')
            
            # Remove rows where required fields are empty
            df = df.dropna(subset=['vendorName', 'contactPerson', 'phone'], how='any')
            
            # Reset index after filtering
            df = df.reset_index(drop=True)
            
        except Exception as e:
            return {"errFlag": 1, "message": f"Could not read Excel sheet: {str(e)}"}

        # Fetch raw materials for dropdown mapping
        with db.engine.connect() as conn:
            raw_materials = conn.execute(text("SELECT id, material_name FROM raw_materials WHERE status=1")).mappings().all()

        raw_material_map = {r['material_name']: r['id'] for r in raw_materials}

        added = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                # Skip row if vendorName is empty (additional safety check)
                if pd.isna(row.get("vendorName")) or not str(row["vendorName"]).strip():
                    continue
                    
                vendorName = str(row["vendorName"]).strip()
                contactPerson = str(row["contactPerson"]).strip()
                email = str(row.get("email", "")).strip() if pd.notna(row.get("email")) else ""
                phone = str(row.get("phone", "")).strip() if pd.notna(row.get("phone")) else ""
                address = str(row.get("address", "")).strip() if pd.notna(row.get("address")) else ""
                city = str(row.get("city", "")).strip() if pd.notna(row.get("city")) else ""
                state = str(row.get("state", "")).strip() if pd.notna(row.get("state")) else ""
                pincode = str(row.get("pincode", "")).strip() if pd.notna(row.get("pincode")) else ""
                gstNo = str(row.get("gstNo", "")).strip() if pd.notna(row.get("gstNo")) else ""
                panNo = str(row.get("panNo", "")).strip() if pd.notna(row.get("panNo")) else ""
                bankName = str(row.get("bankName", "")).strip() if pd.notna(row.get("bankName")) else ""
                accountNo = str(row.get("accountNo", "")).strip() if pd.notna(row.get("accountNo")) else ""
                ifscCode = str(row.get("ifscCode", "")).strip() if pd.notna(row.get("ifscCode")) else ""
                paymentTerms = str(row.get("paymentTerms", "")).strip() if pd.notna(row.get("paymentTerms")) else ""
                creditLimit = float(row.get("creditLimit", 0.0)) if pd.notna(row.get("creditLimit")) else 0.0
                notes = str(row.get("notes", "")).strip() if pd.notna(row.get("notes")) else ""

                # Handle raw materials (comma-separated material names)
                raw_materials_list = []
                raw_materials_str = str(row.get("rawMaterials", "")).strip() if pd.notna(row.get("rawMaterials", "")) else ""
                if raw_materials_str:
                    material_names = [name.strip() for name in raw_materials_str.split(',')]
                    for material_name in material_names:
                        if material_name and material_name in raw_material_map:  # Check if material_name is not empty
                            raw_materials_list.append({"raw_material_id": raw_material_map[material_name]})

                # Validate required fields
                if not vendorName or not contactPerson or not phone:
                    errors.append(f"Row {idx + 2}: Missing required fields (vendorName, contactPerson, or phone)")
                    continue

                # Duplicate check before inserting
                if self.chkDuplicateVendorName(vendorName):
                    errors.append(f"Row {idx + 2}: Vendor '{vendorName}' already exists")
                    continue

                res = self.addVendor(
                    vendorName, contactPerson, email, phone,
                    address, city, state, pincode, gstNo, panNo, bankName, accountNo,
                    ifscCode, paymentTerms, creditLimit, notes, admin_user_id, 
                    None,  # vendorLogoFile skipped in bulk
                    raw_materials_list  # raw materials
                )

                if isinstance(res, dict) and res.get("errFlag") == 1:
                    errors.append(f"Row {idx + 2}: {res['message']}")
                else:
                    added += 1

            except Exception as e:
                errors.append(f"Error in row {idx + 2}: {str(e)}")

        # Prepare response
        if errors:
            error_message = f"{added} vendors uploaded successfully, but {len(errors)} errors occurred: " + "; ".join(errors[:5])  # Show first 5 errors
            return {"errFlag": 1, "message": error_message}
        else:
            return {"errFlag": 0, "message": f"{added} vendors uploaded successfully"}

    def generateBulkUploadTemplate(self):
        wb = Workbook()
        main_ws = wb.active
        main_ws.title = "Vendors"
        ref_ws = wb.create_sheet("Lists")

        # Headers for main data entry sheet
        headers = [
            "vendorName", "contactPerson", "email", "phone", "address", 
            "city", "state", "pincode", "gstNo", "panNo", "bankName", 
            "accountNo", "ifscCode", "paymentTerms", "creditLimit", "notes",
            "rawMaterials"
        ]
        main_ws.append(headers)

        # Fetch dropdown data from DB
        with db.engine.connect() as conn:
            raw_materials = conn.execute(
                text("SELECT id, material_name FROM raw_materials WHERE status=1 ORDER BY material_name")
            ).mappings().all()

        # Fill lookup sheet (Lists)
        ref_ws.cell(row=1, column=1, value="Raw Materials")
        for i, rm in enumerate(raw_materials, start=2):
            ref_ws.cell(row=i, column=1, value=rm["material_name"])

        # Create DataValidation for raw materials dropdown
        dv_raw_materials = DataValidation(
            type="list",
            formula1=f"=Lists!$A$2:$A${len(raw_materials)+1}",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",  # Prevent manual entry
            errorTitle="Invalid Input",
            error="Please select valid raw materials from the dropdown list. For multiple, use comma separation."
        )
        main_ws.add_data_validation(dv_raw_materials)
        dv_raw_materials.add("Q2:Q500")

        # Add instructions note
        main_ws.cell(row=1, column=18, value="Instructions:")
        main_ws.cell(row=2, column=18, value="For rawMaterials: Use comma-separated values from dropdown")
        main_ws.cell(row=3, column=18, value="Example: Material A, Material B, Material C")

        # Protect the "Lists" sheet (optional)
        ref_ws.protection.sheet = True
        ref_ws.protection.enable()

        return send_workbook_response(wb, "bulk_vendor_template.xlsx")
   


# Create a single instance of the class to be used by the routes
vendorObj = VendorClass()