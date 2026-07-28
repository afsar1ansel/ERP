# //ProductSkuClass.py
from db import db
from sqlalchemy.sql import text
from datetime import datetime
import cloudinary.uploader
import os
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection
from helper.files import send_workbook_response

class ProductSkuClass:

    def upload_product_image(self, file):
        """Upload product image to Cloudinary with validation"""
        # Validation: File size (3.5MB limit)
        if len(file.read()) > 3.5 * 1024 * 1024:  
            file.seek(0)  # Reset file pointer
            return {"errFlag": 1, "message": "File size must be less than 3.5MB"}
        
        file.seek(0)  # Reset file pointer after checking size
        
        # Validation: File type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
        filename = secure_filename(file.filename)
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return {"errFlag": 1, "message": "Invalid file type. Allowed: PNG, JPG, JPEG, GIF, BMP, WEBP"}
        try:
            # Upload to Cloudinary
            upload_result = cloudinary.uploader.upload(
                file,
                folder="product_skus",
                transformation=[
                    {'width': 300, 'height': 300, 'crop': 'limit'},
                    {'quality': 'auto'},
                    {'format': 'auto'}
                ]
            )
            return {"errFlag": 0, "url": upload_result['secure_url'], "public_id": upload_result['public_id']}
            
        except Exception as e:
            return {"errFlag": 1, "message": "Upload failed"}

    def delete_product_image(self, public_id):
        """Delete image from Cloudinary"""
        try:
            result = cloudinary.uploader.destroy(public_id)
            if result.get('result') == 'ok':
                return True
            return False
        except:
            return False

    def checkDuplicateProductName(self, productName, productId=None):
        data = {'productName': productName}
        if productId:
            sql = text('SELECT * FROM products_sku WHERE product_name = :productName AND id != :productId')
            data['productId'] = productId
        else:
            sql = text('SELECT * FROM products_sku WHERE product_name = :productName')
        
        with db.engine.connect() as conn:
            res = conn.execute(sql, data)
        return res.mappings().all()

    def addProductSku(self, productName, brandId, productCategoryId, productDescription, productImageFile, rawMaterials, adminUserId, minStockLevel, labourFreightCharge=None):
        # Check duplicate product name
        duplicate_check = self.checkDuplicateProductName(productName)
        if duplicate_check:
            return {"errFlag": 1, "message": "Product name already exists"}
        # Handle product image upload
        product_image_url = ""
        product_image_public_id = ""
        if productImageFile and productImageFile.filename != '':
            upload_result = self.upload_product_image(productImageFile)
            if upload_result["errFlag"] == 1:
                return upload_result
            product_image_url = upload_result["url"]
            product_image_public_id = upload_result["public_id"]
        data = {
            'productName': productName,
            'brandId': brandId,
            'productCategoryId': productCategoryId,
            'minStockLevel': minStockLevel,
            'productDescription': productDescription,
            'productImage': product_image_url,
            'productImagePublicId': product_image_public_id,
            'status': 1,
            'createdAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'createdAdminId': adminUserId,
            'labourFreightCharge': labourFreightCharge
        }

        sql = text('''
            INSERT INTO products_sku (
                product_name, 
                brand_id,
                product_category_id,
                min_stock_level,
                product_description,
                product_image,
                product_image_public_id,
                status,
                created_at,
                created_admin_id,
                labour_freight_charge)
            VALUES (
                :productName,
                :brandId, 
                :productCategoryId,
                :minStockLevel,
                :productDescription,
                :productImage,
                :productImagePublicId,
                :status,
                :createdAt,
                :createdAdminId,
                :labourFreightCharge)
        ''')
        
        try:
            with db.engine.connect() as conn:
                # Insert main product
                result = conn.execute(sql, data)
                product_id = result.lastrowid
                # Insert raw materials consumption
                if rawMaterials:
                    for material in rawMaterials:
                        material_data = {
                            'productSkuId': product_id,
                            'rawMaterialId': material['rawMaterialId'],
                            'quantity': material['quantity'],
                            'unit': material['unit'],
                            'status': 1,
                            'createdAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            'createdAdminId': adminUserId
                        }
                        
                        material_sql = text('''
                            INSERT INTO product_raw_material_consumption (
                                product_sku_id,
                                raw_material_id,
                                quantity,
                                unit,
                                status,
                                created_at,
                                created_admin_id)
                            VALUES (
                                :productSkuId,
                                :rawMaterialId,
                                :quantity,
                                :unit,
                                :status,
                                :createdAt,
                                :createdAdminId)
                        ''')
                        conn.execute(material_sql, material_data)
                
                conn.commit()
            
            return product_id
            
        except Exception as e:
            return {"errFlag": 1, "message": "Error adding product SKU"}

    def updateProductSku(self, productId, productName, brandId, productCategoryId, productDescription, productImageFile, rawMaterials, adminUserId, minStockLevel, labourFreightCharge=None):
        # Check duplicate product name excluding current product
        duplicate_check = self.checkDuplicateProductName(productName, productId)
        if duplicate_check:
            return {"errFlag": 1, "message": "Product name already exists"}

        # Get current product details for image cleanup
        current_product = self.getProductSkuDetails(productId)
        current_image_public_id = current_product[0]['product_image_public_id'] if current_product else None

        # Handle product image upload
        product_image_url = current_product[0]['product_image'] if current_product else ""
        product_image_public_id = current_image_public_id
        
        if productImageFile and productImageFile.filename != '':
            upload_result = self.upload_product_image(productImageFile)
            if upload_result["errFlag"] == 1:
                return upload_result
            product_image_url = upload_result["url"]
            product_image_public_id = upload_result["public_id"]
            
            # Delete old image if it exists
            if current_image_public_id:
                self.delete_product_image(current_image_public_id)

        data = {
            'productId': productId,
            'productName': productName,
            'brandId': brandId,
            'productCategoryId': productCategoryId,
            'minStockLevel': minStockLevel,
            'productDescription': productDescription,
            'productImage': product_image_url,
            'productImagePublicId': product_image_public_id,
            'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'updatedAdminId': adminUserId,
            'labourFreightCharge': labourFreightCharge
        }

        sql = text('''
            UPDATE products_sku 
            SET product_name = :productName, 
                brand_id = :brandId,
                product_category_id = :productCategoryId,
                min_stock_level = :minStockLevel,
                product_description = :productDescription,
                product_image = :productImage,
                product_image_public_id = :productImagePublicId,
                updated_at = :updatedAt,
                updated_admin_id = :updatedAdminId,
                labour_freight_charge = :labourFreightCharge
            WHERE id = :productId
        ''')
        
        try:
            with db.engine.connect() as conn:
                # Update main product
                result = conn.execute(sql, data)
                
                # Delete existing materials and add new ones
                if rawMaterials is not None:
                    # Soft delete existing materials
                    delete_sql = text('''
                        UPDATE product_raw_material_consumption 
                        SET status = 0, updated_at = :updatedAt, updated_admin_id = :adminId
                        WHERE product_sku_id = :productId
                    ''')
                    conn.execute(delete_sql, {
                        'productId': productId,
                        'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        'adminId': adminUserId
                    })
                    
                    # Insert new materials
                    for material in rawMaterials:
                        material_data = {
                            'productSkuId': productId,
                            'rawMaterialId': material['rawMaterialId'],
                            'quantity': material['quantity'],
                            'unit': material['unit'],
                            'status': 1,
                            'createdAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            'createdAdminId': adminUserId
                        }
                        
                        material_sql = text('''
                            INSERT INTO product_raw_material_consumption (
                                product_sku_id,
                                raw_material_id,
                                quantity,
                                unit,
                                status,
                                created_at,
                                created_admin_id)
                            VALUES (
                                :productSkuId,
                                :rawMaterialId,
                                :quantity,
                                :unit,
                                :status,
                                :createdAt,
                                :createdAdminId)
                        ''')
                        conn.execute(material_sql, material_data)
                
                conn.commit()
            
            return result.rowcount
            
        except Exception as e:
            return {"errFlag": 1, "message": "Error updating product SKU"}


    def getAllProductSkus(self):
        """Get all product SKUs with their raw materials data"""
        
        # First get all products
        sql = text('''
            SELECT ps.*, b.brand_name, b.brand_code, pc.product_category_name, ps.product_description, ps.min_stock_level
            FROM products_sku ps 
            LEFT JOIN brands b ON ps.brand_id = b.id 
            LEFT JOIN product_categories pc ON ps.product_category_id = pc.id 
            WHERE ps.status = 1 
            ORDER BY ps.product_name
        ''')
        
        with db.engine.connect() as conn:
            products = conn.execute(sql).mappings().all()
        
        # Get raw materials for each product
        products_with_materials = []
        for product in products:
            product_dict = dict(product)
            
            # Get raw materials for this product
            materials_sql = text('''
                SELECT 
                    prmc.*, 
                    rm.material_name, 
                    rm.material_code
                    
                FROM product_raw_material_consumption prmc 
                LEFT JOIN raw_materials rm ON prmc.raw_material_id = rm.id 
                WHERE prmc.product_sku_id = :productId AND prmc.status = 1
                ORDER BY rm.material_name
            ''')
            with db.engine.connect() as conn:
                materials_result = conn.execute(materials_sql, {'productId': product_dict['id']}).mappings().all()
            
            
            
            # Convert materials to list of dictionaries
            product_dict['rawMaterials'] = [dict(material) for material in materials_result]
            
            # Add material count
            product_dict['material_count'] = len(product_dict['rawMaterials'])
            
            products_with_materials.append(product_dict)
        
        return products_with_materials
    
    def getProductSkuDetails(self, productId):
        sql = text('''
            SELECT ps.*, b.brand_name, pc.product_category_name 
            FROM products_sku ps 
            LEFT JOIN brands b ON ps.brand_id = b.id 
            LEFT JOIN product_categories pc ON ps.product_category_id = pc.id 
            WHERE ps.id = :productId AND ps.status = 1
        ''')
        with db.engine.connect() as conn:
            res = conn.execute(sql, {'productId': productId})
        return res.mappings().all()

    def getProductRawMaterials(self, productId):
        sql = text('''
            SELECT prmc.*, rm.material_name, rm.unit_of_measure 
            FROM product_raw_material_consumption prmc 
            LEFT JOIN raw_materials rm ON prmc.raw_material_id = rm.id 
            WHERE prmc.product_sku_id = :productId AND prmc.status = 1
        ''')
        with db.engine.connect() as conn:
            res = conn.execute(sql, {'productId': productId})
        return res.mappings().all()

    def changeProductSkuStatus(self, productId, status):
        data = {
            'productId': productId,
            'status': status,
            'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        sql = text('UPDATE products_sku SET status = :status, updated_at = :updatedAt WHERE id = :productId')
        
        with db.engine.connect() as conn:
            result = conn.execute(sql, data)
            conn.commit()
        
        return result.rowcount

    
    # =============== BULK UPLOAD ===============
    def bulkUploadProductSkus(self, excel_file, admin_user_id):
    
        try:
            sku_df = pd.read_excel(excel_file, sheet_name="ProductSKU")
            rm_df = pd.read_excel(excel_file, sheet_name="ProductRawMaterials")
        except Exception:
            return {"errFlag": 1, "message": "Invalid or missing Excel sheets"}

        # --- Step 1: Fetch lookup mappings ---
        with db.engine.connect() as conn:
            brands = conn.execute(text("SELECT id, brand_name FROM brands WHERE status=1")).mappings().all()
            categories = conn.execute(text("SELECT id, product_category_name FROM product_categories WHERE status=1")).mappings().all()
            raw_materials = conn.execute(text("SELECT id, material_name FROM raw_materials WHERE status=1")).mappings().all()
            units = conn.execute(text("SELECT unit_name FROM units_of_measurement WHERE status=1")).mappings().all()

        brand_map = {b["brand_name"].strip(): b["id"] for b in brands}
        category_map = {c["product_category_name"].strip(): c["id"] for c in categories}
        raw_map = {r["material_name"].strip(): r["id"] for r in raw_materials}
        valid_units = {u["unit_name"].strip() for u in units}

        inserted_products = 0
        inserted_materials = 0

        with db.engine.begin() as conn:  # Auto commit or rollback
            for idx, row in sku_df.iterrows():
                product_name = str(row.get("productName", "")).strip()
                brand_name = str(row.get("Brand", "")).strip()
                category_name = str(row.get("Category", "")).strip()
                description = str(row.get("productDescription", "")).strip()
                min_stock = int(row.get("minStockLevel", 0))

                if not (product_name and brand_name and category_name):
                    return {"errFlag": 1, "message": f"Missing required fields at row {idx + 2} in ProductSKU"}

                brand_id = brand_map.get(brand_name)
                cat_id = category_map.get(category_name)
                if not brand_id or not cat_id:
                    return {"errFlag": 1, "message": f"Invalid Brand or Category for product '{product_name}'"}

                # --- Check if product exists ---
                existing = conn.execute(
                    text("SELECT id FROM products_sku WHERE product_name = :name AND status=1"),
                    {"name": product_name}
                ).mappings().first()

                if existing:
                    product_id = existing["id"]
                else:
                    # Insert new product
                    insert_sql = text('''
                        INSERT INTO products_sku (
                            product_name, brand_id, product_category_id,
                            min_stock_level, product_description,
                            status, created_at, created_admin_id)
                        VALUES (
                            :productName, :brandId, :categoryId, :minStockLevel,
                            :productDescription, 1, :createdAt, :createdAdminId)
                    ''')
                    result = conn.execute(insert_sql, {
                        "productName": product_name,
                        "brandId": brand_id,
                        "categoryId": cat_id,
                        "minStockLevel": min_stock,
                        "productDescription": description,
                        "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "createdAdminId": admin_user_id
                    })
                    product_id = result.lastrowid
                    inserted_products += 1

                # --- Handle raw materials for this product ---
                product_rm = rm_df[rm_df["productName"].astype(str).str.strip() == product_name]

                for _, rm in product_rm.iterrows():
                    raw_name = str(rm.get("rawMaterial", "")).strip()
                    quantity = float(rm.get("quantity", 0))
                    unit = str(rm.get("unit", "")).strip()

                    if not (raw_name and quantity and unit):
                        continue
                    if raw_name not in raw_map or unit not in valid_units:
                        continue

                    raw_id = raw_map[raw_name]

                    # Check if combination already exists
                    exists_rm = conn.execute(
                        text("""
                            SELECT id FROM product_raw_material_consumption
                            WHERE product_sku_id=:pid AND raw_material_id=:rid AND status=1
                        """),
                        {"pid": product_id, "rid": raw_id}
                    ).mappings().first()

                    if exists_rm:
                        continue  # Skip duplicate

                    # Insert new raw material link
                    conn.execute(text('''
                        INSERT INTO product_raw_material_consumption (
                            product_sku_id, raw_material_id, quantity, unit,
                            status, created_at, created_admin_id)
                        VALUES (
                            :productSkuId, :rawMaterialId, :quantity, :unit,
                            1, :createdAt, :createdAdminId)
                    '''), {
                        "productSkuId": product_id,
                        "rawMaterialId": raw_id,
                        "quantity": quantity,
                        "unit": unit,
                        "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "createdAdminId": admin_user_id
                    })
                    inserted_materials += 1

        return {
            "errFlag": 0,
            "message": f"Upload successful: {inserted_products} new products, {inserted_materials} new raw material links added."
        }


    # =============== GENERATE EXCEL TEMPLATE ===============
    def generateBulkUploadTemplate(self):
        wb = Workbook()

        # --- Product SKU Sheet ---
        sku_ws = wb.active
        sku_ws.title = "ProductSKU"
        sku_headers = ["productName", "Brand", "Category", "minStockLevel", "productDescription"]
        sku_ws.append(sku_headers)

        # --- Product Raw Materials Sheet ---
        rm_ws = wb.create_sheet("ProductRawMaterials")
        rm_headers = ["productName", "rawMaterial", "quantity", "unit"]
        rm_ws.append(rm_headers)

        # --- Reference Sheet ---
        ref_ws = wb.create_sheet("Lists")

        # Fetch lookup data from DB
        with db.engine.connect() as conn:
            brands = conn.execute(text("SELECT brand_name FROM brands WHERE status=1")).mappings().all()
            categories = conn.execute(text("SELECT product_category_name FROM product_categories WHERE status=1")).mappings().all()
            raw_materials = conn.execute(text("SELECT material_name FROM raw_materials WHERE status=1")).mappings().all()
            units = conn.execute(text("SELECT unit_name FROM units_of_measurement WHERE status=1")).mappings().all()

        # Fill Lists sheet
        for i, b in enumerate(brands, start=2):
            ref_ws.cell(row=i, column=1, value=b["brand_name"])
        for i, c in enumerate(categories, start=2):
            ref_ws.cell(row=i, column=2, value=c["product_category_name"])
        for i, r in enumerate(raw_materials, start=2):
            ref_ws.cell(row=i, column=3, value=r["material_name"])
        for i, u in enumerate(units, start=2):
            ref_ws.cell(row=i, column=4, value=u["unit_name"])

        # Helper for strict dropdowns
        def strict_dropdown(formula, error_message):
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Selection",
                error=error_message
            )
            return dv

        # === Dropdowns for ProductSKU sheet ===
        dv_brand = strict_dropdown(
            f"=Lists!$A$2:$A${len(brands)+1}",
            "Please select a valid Brand from the dropdown list."
        )
        sku_ws.add_data_validation(dv_brand)
        dv_brand.add("B2:B500")

        dv_category = strict_dropdown(
            f"=Lists!$B$2:$B${len(categories)+1}",
            "Please select a valid Category from the dropdown list."
        )
        sku_ws.add_data_validation(dv_category)
        dv_category.add("C2:C500")

        # === Dropdowns for ProductRawMaterials sheet ===
        # (1) Product Name dropdown dynamically linked to ProductSKU sheet
        dv_product_name = strict_dropdown(
            "=ProductSKU!$A$2:$A$500",
            "Please select a valid Product Name from the ProductSKU sheet."
        )
        rm_ws.add_data_validation(dv_product_name)
        dv_product_name.add("A2:A1000")

        # (2) Raw Material dropdown from Lists sheet
        dv_raw_mat = strict_dropdown(
            f"=Lists!$C$2:$C${len(raw_materials)+1}",
            "Please select a valid Raw Material from the dropdown list."
        )
        rm_ws.add_data_validation(dv_raw_mat)
        dv_raw_mat.add("B2:B1000")

        # (3) Unit dropdown from Lists sheet
        dv_unit = strict_dropdown(
            f"=Lists!$D$2:$D${len(units)+1}",
            "Please select a valid Unit from the dropdown list."
        )
        rm_ws.add_data_validation(dv_unit)
        dv_unit.add("D2:D1000")

        # Protect Lists sheet so user can't modify dropdown data
        ref_ws.protection.sheet = True
        ref_ws.protection.enable()

        return send_workbook_response(wb, "bulk_product_sku_template.xlsx")
    
# Singleton instance
productSkuObj = ProductSkuClass()