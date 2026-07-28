# Receives from production
# classes/ProductionReceiptsClass.py
from db import db
from sqlalchemy.sql import text
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from flask import send_file
import pandas as pd
from openpyxl.styles.protection import Protection
from helper.files import send_workbook_response

class ProductionReceiptsClass:

    def getProductionBatch(self, batchId):
        sql = text('SELECT * FROM production_batch WHERE id = :batchId')
        with db.engine.connect() as conn:
            res = conn.execute(sql, {'batchId': batchId}).mappings().all()
        return dict(res[0]) if res else None

    def getProductSkuDetails(self, productSkuId):
        sql = text('''
            SELECT ps.*,
                   b.id AS brand_id, b.brand_name,
                   pc.id AS category_id, pc.product_category_name
            FROM products_sku ps
            LEFT JOIN brands b ON ps.brand_id = b.id
            LEFT JOIN product_categories pc ON ps.product_category_id = pc.id
            WHERE ps.id = :productSkuId
        ''')
        with db.engine.connect() as conn:
            res = conn.execute(sql, {'productSkuId': productSkuId}).mappings().all()
        return dict(res[0]) if res else None

    def findFinishedGood(self, productName, brandId, categoryId, conn=None):
        """
        Try to find existing finished_goods row by (product_name, brand_id, product_category_id).
        Accepts optional conn to participate in caller transaction.
        """
        sql = text('''
            SELECT * FROM finished_goods
            WHERE product_name = :productName
              AND (brand_id = :brandId OR (:brandId IS NULL AND brand_id IS NULL))
              AND (product_category_id = :categoryId OR (:categoryId IS NULL AND product_category_id IS NULL))
            LIMIT 1
        ''')
        params = {'productName': productName, 'brandId': brandId, 'categoryId': categoryId}
        if conn:
            res = conn.execute(sql, params).mappings().all()
        else:
            with db.engine.connect() as conn2:
                res = conn2.execute(sql, params).mappings().all()
        return dict(res[0]) if res else None

    def createFinishedGood(self, productName, productImage, productImagePublicId,
                           skuCode, brandId, productCategoryId,
                           stockQty, minLevel, maxLevel, storageLocationId,
                           unitPrice, rawMaterialCost, velocity, minStockLevel, goodsStatus,
                           lastProduced, adminUserId, conn=None):
        """
        Insert new finished_goods row. Accepts optional conn to join transaction.
        Returns inserted fg id or error dict.
        """
        data = {
            'productName': productName,
            'productImage': productImage,
            'productImagePublicId': productImagePublicId,
            'skuCode': skuCode,
            'brandId': brandId,
            'productCategoryId': productCategoryId,
            'stockQty': float(stockQty) if stockQty not in (None, '') else 0,
            'minLevel': float(minStockLevel) if minStockLevel not in (None, '') else None,
            'maxLevel': float(maxLevel) if maxLevel not in (None, '') else None,
            'storageLocationId': int(storageLocationId) if storageLocationId not in (None, '') else None,
            'unitPrice': float(unitPrice) if unitPrice not in (None, "") else 0.00,
            'totalValue': 0.00,
            'rawMaterialCost': float(rawMaterialCost) if rawMaterialCost not in (None, "") else 0.00,
            'velocity': velocity,
            'minStockLevel': None,
            'goodsStatus': goodsStatus or 'in-stock',
            'lastProduced': lastProduced,
            'createdAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'createdAdminId': adminUserId
        }

        try:
            data['totalValue'] = round(data['stockQty'] * data['unitPrice'], 2)
        except Exception:
            data['totalValue'] = 0.00

        insert_sql = text('''
            INSERT INTO finished_goods (
                product_name, product_image, product_image_public_id, sku_code,
                brand_id, product_category_id, stock_qty, min_level, max_level,
                storage_location_id, unit_price, total_value, raw_material_cost,
                velocity,goods_status, last_produced, created_at, created_admin_id, status
            ) VALUES (
                :productName, :productImage, :productImagePublicId, :skuCode,
                :brandId, :productCategoryId, :stockQty, :minLevel, :maxLevel,
                :storageLocationId, :unitPrice, :totalValue, :rawMaterialCost,
                :velocity, :goodsStatus, :lastProduced, :createdAt, :createdAdminId, 1
            )
        ''')

        try:
            if conn:
                res = conn.execute(insert_sql, data)
                fg_id = res.lastrowid
            else:
                with db.engine.connect() as conn2:
                    res = conn2.execute(insert_sql, data)
                    fg_id = res.lastrowid
                    conn2.commit()
            return fg_id
        except Exception as e:
            print("Error in createFinishedGood:", e)
            return {"errFlag": 1, "message": "Error while creating finished good"}

    def updateFinishedGoodStock(self, fgId, addQty, storageLocationId, lastProduced, adminUserId, conn=None):
        """
        Update existing finished_good's stock_qty, total_value, storage_location_id, last_produced.
        Accepts optional conn to join transaction.
        """
        sql_get = text('SELECT stock_qty, unit_price FROM finished_goods WHERE id = :fgId')
        params_get = {'fgId': fgId}

        if conn:
            row = conn.execute(sql_get, params_get).mappings().first()
        else:
            with db.engine.connect() as conn2:
                row = conn2.execute(sql_get, params_get).mappings().first()

        if not row:
            return {"errFlag": 1, "message": "Finished good not found"}

        try:
            current_stock = float(row.get('stock_qty') or 0)
        except Exception:
            current_stock = 0.0
        try:
            unit_price = float(row.get('unit_price') or 0.0)
        except Exception:
            unit_price = 0.0

        new_stock = current_stock + float(addQty)
        new_total_value = round(new_stock * unit_price, 2)

        update_sql = text('''
            UPDATE finished_goods
            SET stock_qty = :newStock,
                total_value = :newTotalValue,
                storage_location_id = :storageLocationId,
                last_produced = :lastProduced,
                updated_at = :updatedAt,
                updated_admin_id = :updatedAdminId
            WHERE id = :fgId
        ''')

        update_params = {
            'fgId': fgId,
            'newStock': new_stock,
            'newTotalValue': new_total_value,
            'storageLocationId': int(storageLocationId) if storageLocationId not in (None, '') else None,
            'lastProduced': lastProduced,
            'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'updatedAdminId': adminUserId
        }

        try:
            if conn:
                res = conn.execute(update_sql, update_params)
            else:
                with db.engine.connect() as conn2:
                    res = conn2.execute(update_sql, update_params)
                    conn2.commit()
            return res.rowcount
        except Exception as e:
            print("Error in updateFinishedGoodStock:", e)
            return {"errFlag": 1, "message": "Error while updating finished good stock"}

    def logProductionReceipt(self, productionBatchId, finishedGoodId, storageLocationId, productSkuId, quantity, notes, adminUserId, conn=None):
        """Log production receipt for a production batch and finished good."""
        insert_sql = text('''
            INSERT INTO production_receipts (
                production_batch_id, finished_goods_id, storage_location_id, product_sku_id, received_qty, notes,
                created_at, created_admin_id
            ) VALUES (
                :productionBatchId, :finishedGoodId, :storageLocationId, :productSkuId, :receivedQty, :notes,
                :createdAt, :createdAdminId
            )'''
        )
        params = {
            'productionBatchId': productionBatchId,
            'finishedGoodId': finishedGoodId,
            'storageLocationId': storageLocationId,
            'productSkuId': productSkuId,
            'receivedQty': quantity,
            'notes': notes,
            'createdAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'createdAdminId': adminUserId
        }   
        try:
            if conn:
                res = conn.execute(insert_sql, params)
            else:
                with db.engine.connect() as conn2:
                    res = conn2.execute(insert_sql, params)
                    conn2.commit()
            return res.lastrowid
        except Exception as e:
            print("Error in logProductionReceipt:", e)
            return {"errFlag": 1, "message": "Error while logging production receipt"}
        
    
    def receiveFromProduction(self, productionBatchId, storageLocationId, quantity, notes, adminUserId):
        """
        Main entry: read production batch -> product sku -> create or update finished_goods and return fg id.
        This version uses helper methods and a single transaction.
        """
        if not productionBatchId:
            return {"errFlag": 1, "message": "productionBatchId is required"}

        batch = self.getProductionBatch(productionBatchId)
        if not batch:
            return {"errFlag": 1, "message": "Production batch not found"}

        productSkuId = batch.get('product_id')
        if not productSkuId:
            return {"errFlag": 1, "message": "production batch has no product associated"}

        sku = self.getProductSkuDetails(productSkuId)
        if not sku:
            return {"errFlag": 1, "message": "Product SKU not found"}

        productName = sku.get('product_name')
        productImage = sku.get('product_image')
        productImagePublicId = sku.get('product_image_public_id')
        skuCode = batch.get('production_code') 
        brandId = sku.get('brand_id')
        productCategoryId = sku.get('category_id')
        unitPrice = sku.get('unit_price') if 'unit_price' in sku else 0.00
        rawMaterialCost = 0.00
        velocity = None
        minStockLevel = sku.get('min_stock_level') if 'min_stock_level' in sku else 0
        goodsStatus = 'in-stock'
        lastProduced = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            # Begin a single transaction and pass conn to helpers
            with db.engine.begin() as conn:
                existing_fg = self.findFinishedGood(productName, brandId, productCategoryId, conn=conn)
                if existing_fg:
                    fg_id = existing_fg['id']
                    # use helper to update stock (participates in same transaction)
                    upd = self.updateFinishedGoodStock(fg_id, quantity, storageLocationId, lastProduced, adminUserId, conn=conn)
                    if isinstance(upd, dict) and upd.get('errFlag') == 1:
                        raise Exception(upd.get('message'))
                else:
                    fg_id = self.createFinishedGood(
                        productName, productImage, productImagePublicId,
                        skuCode, brandId, productCategoryId,
                        quantity, None, None, storageLocationId,
                        unitPrice, rawMaterialCost, velocity, minStockLevel, goodsStatus,
                        lastProduced, adminUserId, conn=conn
                    )
                    if isinstance(fg_id, dict) and fg_id.get('errFlag') == 1:
                        raise Exception(fg_id.get('message'))
                
                
                logIntoReceipts = self.logProductionReceipt(
                    productionBatchId, fg_id, productSkuId, storageLocationId,quantity, notes, adminUserId, conn=conn)

                if isinstance(logIntoReceipts, dict) and logIntoReceipts.get('errFlag') == 1:
                    raise Exception(logIntoReceipts.get('message'))

            return {"errFlag": 0, "message": "Received from production successfully", "fgId": fg_id}
        except Exception as e:
            print("Error in receiveFromProduction:", e)
            return {"errFlag": 1, "message": "Error while receiving from production"}

    def getAllProductionReceipts(self):
        # 1) get all receipts
        sql = text('''
            SELECT pr.*,
                   pb.production_code,
                   ps.id AS sku_id,
                   ps.product_name AS sku_product_name,
                   ps.product_image AS sku_product_image,
                   ps.product_image_public_id AS sku_product_image_public_id,
                   ps.brand_id AS sku_brand_id,
                   ps.product_category_id AS sku_category_id,
                   fg.id AS finished_good_id,
                   fg.product_name AS finished_product_name,
                   fg.product_image AS finished_product_image,
                   fg.product_image_public_id AS finished_product_image_public_id
            FROM production_receipts pr
            LEFT JOIN production_batch pb ON pr.production_batch_id = pb.id
            LEFT JOIN products_sku ps ON pb.product_id = ps.id
            LEFT JOIN finished_goods fg ON ps.id = fg.id OR fg.sku_code = ps.id  -- best-effort; adjust if you have direct mapping
            WHERE pr.status = 1
            ORDER BY pr.created_at DESC
        ''')
        # NOTE: the LEFT JOIN to finished_goods above may need adjustment depending on how you map SKU -> FG.
        # If you have a column products_sku.fg_id or finished_goods.product_sku_id, replace the join accordingly.

        with db.engine.connect() as conn:
            rows = conn.execute(sql).mappings().all()

        # rows already contain receipt-level info; production_receipts are single-line so no items grouping needed
        # Convert to dict list
        return [dict(r) for r in rows]        

    # ----------------- generate Excel template with VLOOKUPs -----------------
    
    def generateBulkUploadTemplate(self):
        """
        Generate Excel template for bulk receiving from production.
        - No pb.status filter (as requested).
        - Unlock entire main sheet area first, then lock only auto-filled/formula columns
        (B: productSkuName, C: brandName, D: categoryName, E: planned_qty, F: completed_qty, G: remaining_qty).
        - Editable columns: A (productionBatchCode), H (quantity), I (storageLocation), J (notes).
        """
        wb = Workbook()
        main_ws = wb.active
        main_ws.title = "ProductionReceipts"

        headers = [
            "productionBatchCode", "productSkuName", "brandName", "categoryName",
            "planned_qty", "completed_qty", "remaining_qty", "quantity", "storageLocation", "notes"
        ]
        main_ws.append(headers)

        ref_ws = wb.create_sheet("Lists")

        # Fetch lookup data (NO pb.status filter)
        with db.engine.connect() as conn:
            batches = conn.execute(text(
                """SELECT pb.id, pb.production_code, pb.planned_qty, pb.completed_qty,
                        ps.product_name, b.brand_name, pc.product_category_name
                FROM production_batch pb
                LEFT JOIN products_sku ps ON pb.product_id = ps.id
                LEFT JOIN brands b ON ps.brand_id = b.id
                LEFT JOIN product_categories pc ON ps.product_category_id = pc.id
                ORDER BY pb.production_code
                """
            )).mappings().all()

            storage_locations = conn.execute(text(
                "SELECT id, location_label FROM storage_locations WHERE status=1 ORDER BY location_label"
            )).mappings().all()

        # Fill Lists: columns A..F for batch mapping (A production_code, B product_name, C brand, D category, E planned, F completed)
        for i, b in enumerate(batches, start=2):
            ref_ws.cell(row=i, column=1, value=b.get("production_code") or "")
            ref_ws.cell(row=i, column=2, value=b.get("product_name") or "")
            ref_ws.cell(row=i, column=3, value=b.get("brand_name") or "")
            ref_ws.cell(row=i, column=4, value=b.get("product_category_name") or "")
            ref_ws.cell(row=i, column=5, value=(float(b.get("planned_qty")) if b.get("planned_qty") not in (None, "") else 0))
            ref_ws.cell(row=i, column=6, value=(float(b.get("completed_qty")) if b.get("completed_qty") not in (None, "") else 0))

        # Put storage locations into column H (col index 8)
        start_loc_col = 8
        for i, loc in enumerate(storage_locations, start=2):
            ref_ws.cell(row=i, column=start_loc_col, value=loc.get("location_label") or "")

        batch_count = max(1, len(batches))
        loc_count = max(1, len(storage_locations))
        list_last_row = max(2, batch_count + 1)
        loc_last_row = max(2, loc_count + 1)

        # Strict dropdown helper
        def strict_dropdown(formula, error_message):
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Selection",
                error=error_message
            )
            return dv

        # Dropdown for productionBatchCode in main sheet (A)
        dv_batch = strict_dropdown(f"=Lists!$A$2:$A${list_last_row}", "Please select a valid Production Batch from the list")
        main_ws.add_data_validation(dv_batch)
        dv_batch.add("A2:A500")

        # Dropdown for storageLocation in main sheet (I) referencing Lists col H
        loc_col_letter = chr(64 + start_loc_col)  # 'H'
        dv_loc = strict_dropdown(f"=Lists!${loc_col_letter}$2:${loc_col_letter}${loc_last_row}", "Please select a valid Storage Location")
        main_ws.add_data_validation(dv_loc)
        dv_loc.add("I2:I500")

        # VLOOKUP range for formulas: Lists!$A$2:$F$N
        vlookup_range = f"Lists!$A$2:$F${list_last_row}"

        # PROTECTION STRATEGY:
        # 1) First, unlock the entire main area we'll use (columns A..J, rows 2..500) -> so by default editable.
        # 2) Then lock only the formula/auto columns B..G.
        locked = Protection(locked=True)
        unlocked = Protection(locked=False)

        # Step 1: unlock main usable area (headers row is left as normal)
        for r in range(2, 501):
            for col_idx in range(1, 11):  # A..J
                cell = main_ws.cell(row=r, column=col_idx)
                cell.protection = unlocked

        # Step 2: write formulas into the auto columns and lock them
        for r in range(2, 501):
            a_ref = f"$A{r}"
            # productSkuName (B)
            main_ws[f"B{r}"].value = f'=IF({a_ref}="","",IFERROR(VLOOKUP({a_ref},{vlookup_range},2,FALSE),""))'
            # brandName (C)
            main_ws[f"C{r}"].value = f'=IF({a_ref}="","",IFERROR(VLOOKUP({a_ref},{vlookup_range},3,FALSE),""))'
            # categoryName (D)
            main_ws[f"D{r}"].value = f'=IF({a_ref}="","",IFERROR(VLOOKUP({a_ref},{vlookup_range},4,FALSE),""))'
            # planned_qty (E)
            main_ws[f"E{r}"].value = f'=IF({a_ref}="","",IFERROR(VLOOKUP({a_ref},{vlookup_range},5,FALSE),""))'
            # completed_qty (F)
            main_ws[f"F{r}"].value = f'=IF({a_ref}="","",IFERROR(VLOOKUP({a_ref},{vlookup_range},6,FALSE),""))'
            # remaining_qty (G) = planned - completed
            main_ws[f"G{r}"].value = f'=IF(OR(E{r}="",F{r}=""),"",E{r}-F{r})'

            # Now lock these auto columns for this row
            for col_letter in ["B", "C", "D", "E", "F", "G"]:
                main_ws[f"{col_letter}{r}"].protection = locked

            # Ensure editable columns remain unlocked (A, H, I, J)
            for col_letter in ["A", "H", "I", "J"]:
                main_ws[f"{col_letter}{r}"].protection = unlocked

        # Protect MAIN sheet but allow selecting unlocked cells
        main_ws.protection.sheet = True
        main_ws.protection.enable()
        main_ws.protection.selectLockedCells = False
        main_ws.protection.selectUnlockedCells = True

        # Protect Lists sheet fully (no edits). Users can view it but not change it.
        ref_ws.protection.sheet = True
        ref_ws.protection.enable()
        ref_ws.protection.selectLockedCells = False
        ref_ws.protection.selectUnlockedCells = False

        return send_workbook_response(wb, "bulk_production_receipts_template.xlsx")


    # ----------------- bulk upload processor -----------------
    def bulkUploadProductionReceipts(self, excel_file, admin_user_id, continue_on_error=True):
        """
        Read uploaded Excel (ProductionReceipts sheet). For each row:
        - Validate batch exists
        - Validate quantity <= remaining (planned - completed)
        - Call receiveFromProduction() for each row (reuses existing logic)
        Returns summary with successes and failures.
        If continue_on_error is False -> stops and returns the first error (atomic)
        """
        try:
            df = pd.read_excel(excel_file, sheet_name="ProductionReceipts")
        except Exception as e:
            print("bulkUpload read error:", e)
            return {"errFlag": 1, "message": "Could not read 'ProductionReceipts' sheet from Excel"}

        # Normalize column names that user might have trimmed/cased differently
        df_columns = {c.strip(): c for c in df.columns}
        # Expected columns: productionBatchCode, quantity, storageLocation, notes
        # But template also contains auto-filled columns; we only need productionBatchCode, quantity, storageLocation, notes
        def col(name):
            return df_columns.get(name, None)

        required_col = col("productionBatchCode")
        qty_col = col("quantity")
        loc_col = col("storageLocation")
        notes_col = col("notes")

        if required_col is None or qty_col is None:
            return {"errFlag": 1, "message": "Excel missing required columns: productionBatchCode and/or quantity"}

        successes = []
        failures = []

        # Preload batch map and locations for faster lookup
        with db.engine.connect() as conn:
            batches = conn.execute(text(
                "SELECT id, production_code, planned_qty, completed_qty, product_id FROM production_batch "
            )).mappings().all()
            batch_map = {b["production_code"]: b for b in batches}

            locs = conn.execute(text("SELECT id, location_label FROM storage_locations WHERE status=1")).mappings().all()
            loc_map = {l["location_label"]: l["id"] for l in locs}

        # Process rows
        for idx, row in df.iterrows():
            row_num = idx + 2
            try:
                batch_code = str(row.get(required_col)).strip() if pd.notna(row.get(required_col)) else ""
                if batch_code == "" or batch_code.lower() == "nan":
                    failures.append({"row": row_num, "error": "Empty productionBatchCode"})
                    if not continue_on_error:
                        return {"errFlag": 1, "message": f"Row {row_num}: productionBatchCode required"}
                    else:
                        continue

                # parse quantity
                raw_qty = row.get(qty_col)
                if pd.isna(raw_qty) or raw_qty == "":
                    failures.append({"row": row_num, "error": "Quantity missing"})
                    if not continue_on_error:
                        return {"errFlag": 1, "message": f"Row {row_num}: quantity required"}
                    else:
                        continue
                try:
                    quantity = float(raw_qty)
                except Exception:
                    failures.append({"row": row_num, "error": "Invalid quantity format"})
                    if not continue_on_error:
                        return {"errFlag": 1, "message": f"Row {row_num}: invalid quantity"}
                    else:
                        continue

                if quantity <= 0:
                    failures.append({"row": row_num, "error": "Quantity must be > 0"})
                    if not continue_on_error:
                        return {"errFlag": 1, "message": f"Row {row_num}: invalid quantity"}
                    else:
                        continue

                batch = batch_map.get(batch_code)
                if not batch:
                    failures.append({"row": row_num, "error": f"Production batch '{batch_code}' not found"})
                    if not continue_on_error:
                        return {"errFlag": 1, "message": f"Row {row_num}: batch not found"}
                    else:
                        continue

                planned = float(batch.get("planned_qty") or 0)
                completed = float(batch.get("completed_qty") or 0)
                remaining = planned - completed

                if quantity > remaining:
                    failures.append({"row": row_num, "error": f"Quantity {quantity} exceeds remaining planned qty {remaining} for batch {batch_code}"})
                    if not continue_on_error:
                        return {"errFlag": 1, "message": f"Row {row_num}: quantity exceeds remaining"}
                    else:
                        continue

                # storage location mapping (optional)
                storage_label = str(row.get(loc_col)).strip() if loc_col and pd.notna(row.get(loc_col)) else None
                storage_id = None
                if storage_label:
                    storage_id = loc_map.get(storage_label)
                    if storage_id is None:
                        failures.append({"row": row_num, "error": f"Storage location '{storage_label}' not found"})
                        if not continue_on_error:
                            return {"errFlag": 1, "message": f"Row {row_num}: invalid storage location"}
                        else:
                            continue

                notes = str(row.get(notes_col)).strip() if notes_col and pd.notna(row.get(notes_col)) else ""

                # Call existing logic (per-row transaction inside method)
                res = self.receiveFromProduction(batch["id"], storage_id, quantity, notes, admin_user_id)
                if isinstance(res, dict) and res.get("errFlag") == 1:
                    failures.append({"row": row_num, "error": res.get("message")})
                    if not continue_on_error:
                        return {"errFlag": 1, "message": f"Row {row_num}: {res.get('message')}"}
                    else:
                        continue

                successes.append({"row": row_num, "batch": batch_code, "received": quantity, "fgId": res.get("fgId") if isinstance(res, dict) else None})

            except Exception as e:
                print(f"Row {row_num} exception:", e)
                failures.append({"row": row_num, "error": str(e)})
                if not continue_on_error:
                    return {"errFlag": 1, "message": f"Row {row_num}: {str(e)}"}

        return {
            "errFlag": 0,
            "message": f"Bulk processing completed. Success: {len(successes)}, Failed: {len(failures)}",
            "successes": successes,
            "failures": failures
        }


# singleton instance
productionReceiptsObj = ProductionReceiptsClass()
