from db import db
from sqlalchemy.sql import text
from datetime import datetime
import cloudinary.uploader
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from flask import send_file
from io import BytesIO
import pandas as pd
from helper.files import send_workbook_response




class RawMaterialClass:

    def upload_raw_material_image(self, file):
        """Upload raw material image to Cloudinary with validation"""
        
        # Validation: File size (2MB limit)
        if len(file.read()) > 2 * 1024 * 1024:  # 2MB in bytes
            file.seek(0)  # Reset file pointer
            return {"errFlag": 1, "message": "File size must be less than 2MB"}
        
        file.seek(0)  # Reset file pointer after checking size
        
        # Validation: File type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
        filename = secure_filename(file.filename)
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            return {"errFlag": 1, "message": "Invalid file type. Allowed: PNG, JPG, JPEG, GIF, BMP, WEBP"}
        
        try:
            # Upload to Cloudinary
            upload_result = cloudinary.uploader.upload(
                file,
                folder="raw_materials",
                transformation=[
                    {'width': 300, 'height': 300, 'crop': 'limit'},
                    {'quality': 'auto'},
                    {'format': 'auto'}
                ]
            )
            
            return {"errFlag": 0, "url": upload_result['secure_url'], "public_id": upload_result['public_id']}
            
        except Exception as e:
            return {"errFlag": 1, "message": "Upload failed"}

    def delete_raw_material_image(self, public_id):
        """Delete image from Cloudinary"""
        try:
            result = cloudinary.uploader.destroy(public_id)
            if result.get('result') == 'ok':
                return True
            return False
        except:
            return False

    def checkDuplicateMaterialCode(self, materialCode, materialId=None):
        data = {'materialCode': materialCode}

        if materialId:
            sql = text('SELECT * FROM raw_materials WHERE material_code = :materialCode AND id != :materialId AND status = 1')
            data['materialId'] = materialId
        else:
            sql = text('SELECT * FROM raw_materials WHERE material_code = :materialCode AND status = 1')

        with db.engine.connect() as conn:
            res = conn.execute(sql, data)
        return res.mappings().all()

    def checkDuplicateMaterialName(self, materialName, materialId=None):
        data = {'materialName': materialName}

        if materialId:
            sql = text('SELECT * FROM raw_materials WHERE material_name = :materialName AND status = 1 AND id != :materialId')
            data['materialId'] = materialId
        else:
            sql = text('SELECT * FROM raw_materials WHERE material_name = :materialName AND status = 1')
        
        with db.engine.connect() as conn:
            res = conn.execute(sql, data)
        return res.mappings().all()

    def addRawMaterial(self, materialCode, materialName, materialDescription, rawMaterialCategoryId, vendorIds, specification, stockQty, minStockLevel, maxStockLevel, unitOfMeasure, storageLocationId, unitCost, materialImageFile, adminUserId):
        # Check duplicate material code
        duplicate_code = self.checkDuplicateMaterialCode(materialCode)
        if duplicate_code:
            return {"errFlag": 1, "message": "Material code already exists"}

        # Check duplicate material name
        duplicate_name = self.checkDuplicateMaterialName(materialName)
        if duplicate_name:
            return {"errFlag": 1, "message": "Material name already exists"}

        # Calculate total value
        total_value = float(stockQty) * float(unitCost)

        # Handle material image upload
        material_image_url = ""
        material_image_public_id = ""
        
        if materialImageFile and materialImageFile.filename != '':
            upload_result = self.upload_raw_material_image(materialImageFile)
            if upload_result["errFlag"] == 1:
                return upload_result
            material_image_url = upload_result["url"]
            material_image_public_id = upload_result["public_id"]

        # Determine stock status
        stock_status = "in-stock"
        if stockQty == 0:
            stock_status = "out-of-stock"
        elif minStockLevel and stockQty <= minStockLevel:
            stock_status = "low-stock"

        data = {
            'materialCode': materialCode,
            'materialName': materialName,
            'materialDescription': materialDescription,
            'rawMaterialCategoryId': rawMaterialCategoryId,
            'specification': specification,
            'stockQty': stockQty,
            'minStockLevel': minStockLevel,
            'maxStockLevel': maxStockLevel,
            'unitOfMeasure': unitOfMeasure,
            'storageLocationId': storageLocationId,
            'unitCost': unitCost,
            'totalValue': total_value,
            'stockStatus': stock_status,
            'materialImage': material_image_url,
            'materialImagePublicId': material_image_public_id,
            'status': 1,
            'createdAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'createdAdminId': adminUserId
        }

        sql = text('''
            INSERT INTO raw_materials (
                material_code, 
                material_name,
                material_description,
                raw_material_category_id,
                specification,
                stock_qty,
                min_stock_level,
                max_stock_level,
                unit_of_measure,
                storage_location_id,
                unit_cost,
                total_value,
                stock_status,
                raw_material_image,
                raw_material_image_public_id,
                status,
                created_at,
                created_admin_id)
            VALUES (
                :materialCode,
                :materialName, 
                :materialDescription,
                :rawMaterialCategoryId,
                :specification,
                :stockQty,
                :minStockLevel,
                :maxStockLevel,
                :unitOfMeasure,
                :storageLocationId,
                :unitCost,
                :totalValue,
                :stockStatus,
                :materialImage,
                :materialImagePublicId,
                :status,
                :createdAt,
                :createdAdminId)
        ''')
        
        try:
            with db.engine.connect() as conn:
                result = conn.execute(sql, data)
                new_material_id = result.lastrowid
                
                # Insert vendors
                if vendorIds and isinstance(vendorIds, list):
                    vendor_values = []
                    for v_id in vendorIds:
                        if v_id: 
                            vendor_values.append({'rm_id': new_material_id, 'v_id': v_id})
                    
                    if vendor_values:
                        conn.execute(text("INSERT INTO raw_material_vendors (raw_material_id, vendor_id) VALUES (:rm_id, :v_id)"), vendor_values)

                conn.commit()
            return result.rowcount
        except Exception as e:
            print("Error adding raw material:", e)
            return {"errFlag": 1, "message": "Database error while adding raw material"}

    def updateRawMaterial(self, materialId, materialCode, materialName, materialDescription, rawMaterialCategoryId, vendorIds, specification, stockQty, minStockLevel, maxStockLevel, unitOfMeasure, storageLocationId, unitCost, materialImageFile, adminUserId):
        # Check duplicate material code excluding current material
        duplicate_code = self.checkDuplicateMaterialCode(materialCode, materialId)
        if duplicate_code:
            return {"errFlag": 1, "message": "Material code already exists"}

        # Check duplicate material name excluding current material
        duplicate_name = self.checkDuplicateMaterialName(materialName, materialId)
        if duplicate_name:
            return {"errFlag": 1, "message": "Material name already exists"}

        # Calculate total value
        total_value = float(stockQty) * float(unitCost)

        # Get current material details for image cleanup
        current_material = self.getRawMaterialDetails(materialId)
        current_image_public_id = current_material[0]['raw_material_image_public_id'] if current_material and 'raw_material_image_public_id' in current_material[0] else None

        # Handle material image upload
        material_image_url = current_material[0]['raw_material_image'] if current_material and 'raw_material_image' in current_material[0] else ""
        material_image_public_id = current_image_public_id
        
        if materialImageFile and materialImageFile.filename != '':
            upload_result = self.upload_raw_material_image(materialImageFile)
            if upload_result["errFlag"] == 1:
                return upload_result
            material_image_url = upload_result["url"]
            material_image_public_id = upload_result["public_id"]
            
            # Delete old image if it exists
            if current_image_public_id:
                self.delete_raw_material_image(current_image_public_id)

        # Determine stock status
        stock_status = "in-stock"
        if stockQty == 0:
            stock_status = "out-of-stock"
        elif minStockLevel and stockQty <= minStockLevel:
            stock_status = "low-stock"

        data = {
            'materialId': materialId,
            'materialCode': materialCode,
            'materialName': materialName,
            'materialDescription': materialDescription,
            'rawMaterialCategoryId': rawMaterialCategoryId,
            'specification': specification,
            'stockQty': stockQty,
            'minStockLevel': minStockLevel,
            'maxStockLevel': maxStockLevel,
            'unitOfMeasure': unitOfMeasure,
            'storageLocationId': storageLocationId,
            'unitCost': unitCost,
            'totalValue': total_value,
            'stockStatus': stock_status,
            'materialImage': material_image_url,
            'materialImagePublicId': material_image_public_id,
            'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'updatedAdminId': adminUserId
        }

        sql = text('''
            UPDATE raw_materials 
            SET material_code = :materialCode, 
                material_name = :materialName,
                material_description = :materialDescription,
                raw_material_category_id = :rawMaterialCategoryId,
                specification = :specification,
                stock_qty = :stockQty,
                min_stock_level = :minStockLevel,
                max_stock_level = :maxStockLevel,
                unit_of_measure = :unitOfMeasure,
                storage_location_id = :storageLocationId,
                unit_cost = :unitCost,
                total_value = :totalValue,
                stock_status = :stockStatus,
                raw_material_image = :materialImage,
                raw_material_image_public_id = :materialImagePublicId,
                updated_at = :updatedAt,
                updated_admin_id = :updatedAdminId
            WHERE id = :materialId
        ''')
        
        try:
            with db.engine.connect() as conn:
                result = conn.execute(sql, data)
                
                # Update vendors
                conn.execute(text("DELETE FROM raw_material_vendors WHERE raw_material_id = :materialId"), {'materialId': materialId})
                
                if vendorIds and isinstance(vendorIds, list):
                    vendor_values = []
                    for v_id in vendorIds:
                        if v_id:
                            vendor_values.append({'rm_id': materialId, 'v_id': v_id})
                    
                    if vendor_values:
                        conn.execute(text("INSERT INTO raw_material_vendors (raw_material_id, vendor_id) VALUES (:rm_id, :v_id)"), vendor_values)
                
                conn.commit()
            return 1
        except Exception as e:
            print("Error updating raw material:", e)
            return {"errFlag": 1, "message": "Database error while updating raw material"}

    def getAllRawMaterials(self):
        sql = text('''
            SELECT rm.*,
                rmc.category_name,
                sl.location_label, 
                sl.capacity,
                GROUP_CONCAT(v.vendor_name SEPARATOR ', ') as vendor_names,
                GROUP_CONCAT(v.id SEPARATOR ',') as vendor_ids
            FROM raw_materials rm 
            LEFT JOIN raw_material_categories rmc ON rm.raw_material_category_id = rmc.id 
            LEFT JOIN raw_material_vendors rmv ON rm.id = rmv.raw_material_id
            LEFT JOIN vendors v ON rmv.vendor_id = v.id 
            LEFT JOIN storage_locations sl ON rm.storage_location_id = sl.id
            WHERE rm.status = 1 
            GROUP BY rm.id
            ORDER BY rm.material_name
        ''')
        with db.engine.connect() as conn:
            res = conn.execute(sql)
        return res.mappings().all()

    def getRawMaterialDetails(self, materialId):
        sql = text('''
            SELECT rm.*, rmc.category_name, sl.*,
                   GROUP_CONCAT(v.vendor_name SEPARATOR ', ') as vendor_names,
                   GROUP_CONCAT(v.id SEPARATOR ',') as vendor_ids
            FROM raw_materials rm 
            LEFT JOIN raw_material_categories rmc ON rm.raw_material_category_id = rmc.id 
            LEFT JOIN raw_material_vendors rmv ON rm.id = rmv.raw_material_id
            LEFT JOIN vendors v ON rmv.vendor_id = v.id 
            LEFT JOIN storage_locations sl ON rm.storage_location_id = sl.id
            WHERE rm.id = :materialId AND rm.status = 1
            GROUP BY rm.id
        ''')
        with db.engine.connect() as conn:
            res = conn.execute(sql, {'materialId': materialId})
        return res.mappings().all()

    def changeRawMaterialStatus(self, materialId, status):
        data = {
            'materialId': materialId,
            'status': status,
            'updatedAt': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        sql = text('UPDATE raw_materials SET status = :status, updated_at = :updatedAt WHERE id = :materialId')
        
        with db.engine.connect() as conn:
            result = conn.execute(sql, data)
            conn.commit()
        
        return result.rowcount

    def getMaterialsByCategories(self, categoryIds):
        """
        Get all raw materials for a list of category IDs.
        Returns a list of rows with material + category info.
        """
        # sanitize input - expect a list of ints
        if not categoryIds or not isinstance(categoryIds, (list, tuple)):
            return []

        # ensure all are ints
        try:
            ids = [int(i) for i in categoryIds]
        except Exception:
            return []

        # build safe parameterized IN clause (:id0, :id1, ...)
        placeholders = ','.join([f':id{i}' for i in range(len(ids))])
        sql = text(f'''
            SELECT 
                rm.*,
                rmc.category_name,
                rmc.id AS category_id,
                GROUP_CONCAT(v.vendor_name SEPARATOR ', ') as vendor_names
            FROM raw_materials rm
            LEFT JOIN raw_material_categories rmc ON rm.raw_material_category_id = rmc.id
            LEFT JOIN raw_material_vendors rmv ON rm.id = rmv.raw_material_id
            LEFT JOIN vendors v ON rmv.vendor_id = v.id
            WHERE rm.raw_material_category_id IN ({placeholders})
              AND rm.status = 1
            GROUP BY rm.id
            ORDER BY rmc.category_name, rm.material_name
        ''')
        # build params dict
        params = {f'id{i}': ids[i] for i in range(len(ids))}

        try:
            with db.engine.connect() as conn:
                responseData = conn.execute(sql, params)
                rows = responseData.mappings().all()
            return rows
        except Exception as e:
            return []


    ###-----------------Bulk raw material upload-------------- ###
    
    def bulkUploadRawMaterials(self, excel_file, admin_user_id):
        try:
            df = pd.read_excel(excel_file, sheet_name="RawMaterials")
        except Exception:
            return {"errFlag": 1, "message": "Could not read Excel sheet"}

        # Fetch category, vendor, storage location mapping
        with db.engine.connect() as conn:
            categories = conn.execute(text("SELECT id, category_name FROM raw_material_categories WHERE status=1")).mappings().all()
            vendors = conn.execute(text("SELECT id, vendor_name FROM vendors WHERE status=1")).mappings().all()
            locations = conn.execute(text("SELECT id, location_label FROM storage_locations WHERE status=1")).mappings().all()

        category_map = {r['category_name']: r['id'] for r in categories}
        vendor_map = {r['vendor_name']: r['id'] for r in vendors}
        location_map = {r['location_label']: r['id'] for r in locations}

        added = 0
        for idx, row in df.iterrows():
            try:
                materialCode = str(row["materialCode"]).strip()
                materialName = str(row["materialName"]).strip()
                materialDescription = str(row.get("materialDescription", "")).strip()
                rawMaterialCategoryId = category_map.get(str(row.get("Category", "")).strip())
                
                # Handle multiple vendors
                vendor_names = str(row.get("Vendor", "")).strip()
                vendorIds = []
                if vendor_names and vendor_names.lower() != 'nan':
                     for v_name in vendor_names.split(','):
                        v_id = vendor_map.get(v_name.strip())
                        if v_id:
                            vendorIds.append(v_id)
                
                specification = str(row.get("Specification", "")).strip()
                stockQty = int(row.get("stockQty", 0))
                minStockLevel = int(row.get("minStockLevel")) if pd.notna(row.get("minStockLevel")) else None
                maxStockLevel = int(row.get("maxStockLevel")) if pd.notna(row.get("maxStockLevel")) else None
                unitOfMeasure = str(row.get("unitOfMeasure")).strip()
                storageLocationId = location_map.get(str(row.get("StorageLocation", "")).strip()) if row.get("StorageLocation") else None
                unitCost = float(row.get("unitCost", 0.0))

                # Duplicate checks before inserting
                if self.checkDuplicateMaterialCode(materialCode):
                    print(f"Skipping duplicate materialCode: {materialCode}")
                    continue
                if self.checkDuplicateMaterialName(materialName):
                    print(f"Skipping duplicate materialName: {materialName}")
                    continue
                
                if not materialCode or not materialName or not rawMaterialCategoryId or not unitOfMeasure:
                    return {"errFlag": 1, "message": f"Missing required fields at row {idx + 2}"}

                res = self.addRawMaterial(
                    materialCode, materialName, materialDescription,
                    rawMaterialCategoryId, vendorIds, specification,
                    stockQty, minStockLevel, maxStockLevel,
                    unitOfMeasure, storageLocationId, unitCost,
                    None,  # imageFile skipped in bulk
                    admin_user_id
                )

                if isinstance(res, dict) and res.get("errFlag") == 1:
                    return {"errFlag": 1, "message": f"Row {idx + 2} error: {res['message']}"}
                added += 1

            except Exception as e:
                return {"errFlag": 1, "message": f"Error in row {idx + 2}: {str(e)}"}

        return {"errFlag": 0, "message": f"{added} raw materials uploaded successfully"}


    def generateBulkUploadTemplate(self):
        wb = Workbook()
        main_ws = wb.active
        main_ws.title = "RawMaterials"
        ref_ws = wb.create_sheet("Lists")

        # Headers for main data entry sheet
        headers = [
            "materialCode", "materialName", "materialDescription",
            "Category", "Vendor", "Specification", "stockQty",
            "minStockLevel", "maxStockLevel", "unitOfMeasure",
            "StorageLocation", "unitCost"
        ]
        main_ws.append(headers)

        # Fetch dropdown data from DB
        with db.engine.connect() as conn:
            categories = conn.execute(
                text("SELECT id, category_name FROM raw_material_categories WHERE status=1")
            ).mappings().all()
            vendors = conn.execute(
                text("SELECT id, vendor_name FROM vendors WHERE status=1")
            ).mappings().all()
            locations = conn.execute(
                text("SELECT id, location_label FROM storage_locations WHERE status=1")
            ).mappings().all()
            units = conn.execute(
                text("SELECT id, unit_name FROM units_of_measurement WHERE status=1")
            ).mappings().all()

        # Fill lookup sheet (Lists)
        for i, cat in enumerate(categories, start=2):
            ref_ws.cell(row=i, column=1, value=cat["category_name"])
        for i, ven in enumerate(vendors, start=2):
            ref_ws.cell(row=i, column=2, value=ven["vendor_name"])
        for i, loc in enumerate(locations, start=2):
            ref_ws.cell(row=i, column=3, value=loc["location_label"])
        for i, unit in enumerate(units, start=2):
            ref_ws.cell(row=i, column=4, value=unit["unit_name"])

        # Utility function for strict dropdowns
        def create_strict_dropdown(formula, error_message):
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",  # Prevent manual entry
                errorTitle="Invalid Input",
                error=error_message
            )
            return dv

        # Category dropdown (D column)
        dv_cat = create_strict_dropdown(
            f"=Lists!$A$2:$A${len(categories)+1}",
            "Please select a valid Category from the dropdown list."
        )
        main_ws.add_data_validation(dv_cat)
        dv_cat.add("D2:D500")

        # Vendor dropdown (E column) - Allow multiple (loose validation)
        dv_vendor = DataValidation(
            type="list",
            formula1=f"=Lists!$B$2:$B${len(vendors)+1}",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="warning", 
            errorTitle="Input Warning",
            error="Select from list. For multiple vendors, type names separated by commas (e.g. Vendor A, Vendor B)."
        )
        main_ws.add_data_validation(dv_vendor)
        dv_vendor.add("E2:E500")

        # Unit of Measure dropdown (J column)
        dv_unit = create_strict_dropdown(
            f"=Lists!$D$2:$D${len(units)+1}",
            "Please select a valid Unit of Measure from the dropdown list."
        )
        main_ws.add_data_validation(dv_unit)
        dv_unit.add("J2:J500")

        # Storage Location dropdown (K column)
        dv_loc = create_strict_dropdown(
            f"=Lists!$C$2:$C${len(locations)+1}",
            "Please select a valid Storage Location from the dropdown list."
        )
        main_ws.add_data_validation(dv_loc)
        dv_loc.add("K2:K500")

        # Protect the "Lists" sheet (optional)
        ref_ws.protection.sheet = True
        ref_ws.protection.enable()

        
        return send_workbook_response(wb, "bulk_raw_material_template.xlsx")



# Singleton instance
rawMaterialObj = RawMaterialClass()