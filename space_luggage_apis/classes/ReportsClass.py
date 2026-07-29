# ReportClass.py
from db import db
from sqlalchemy.sql import text
from datetime import datetime
import requests
import csv
import io
import pandas as pd
from flask import Response
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import os
from openpyxl.styles import PatternFill
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Image, Spacer 
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.colors import HexColor
from collections import defaultdict
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.platypus.flowables import HRFlowable

class ReportsClass:
    
    # Unified Report Generators
    def _generate_csv(self, data, columns, filename_prefix, start_date, end_date):
        """Generate CSV report from data"""
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header row
            writer.writerow([col['header'] for col in columns])
            
            # Write data rows
            for row in data:
                writer.writerow([
                    col['formatter'](row[col['key']]) if 'formatter' in col else str(row[col['key']] or '')
                    for col in columns
                ])
            
            csv_output = output.getvalue()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{start_date}_to_{end_date}_{timestamp}.csv"
            
            return Response(
                csv_output,
                mimetype="text/csv",
                headers={
                    "Content-Disposition": f"attachment;filename={filename}",
                    "Content-Type": "text/csv; charset=utf-8"
                }
            )
        except Exception as e:
            print("Error generating CSV:", e)
            return {"errFlag": 1, "message": f"Error generating CSV: {str(e)}"}
    
    def _generate_excel(self, data, columns, filename_prefix, start_date, end_date, row_styler=None):
        """Generate Excel report from data with optional row styling"""
        try:
            # Create DataFrame
            df_data = []
            for row in data:
                df_data.append([
                    col['formatter'](row[col['key']]) if 'formatter' in col else row[col['key']]
                    for col in columns
                ])
            
            df = pd.DataFrame(df_data, columns=[col['header'] for col in columns])
            
            # Create Excel file in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Report', index=False)
                
                # Apply row styling if provided
                if row_styler:
                    worksheet = writer.sheets['Report']
                    row_styler(worksheet, data, columns)
                
                # Auto-adjust column widths
                worksheet = writer.sheets['Report']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{start_date}_to_{end_date}_{timestamp}.xlsx"
            
            return Response(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment;filename={filename}",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
            )
        except Exception as e:
            print("Error generating Excel:", e)
            return {"errFlag": 1, "message": f"Error generating Excel: {str(e)}"}

            # pdf reports functions 

    def _calculate_column_widths_advanced(self, columns, table_data, available_width, styles):
        """More sophisticated column width calculation"""
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # Try to use Helvetica for width calculation
        try:
            font_name = 'Helvetica'
            bold_font_name = 'Helvetica-Bold'
            
            # Column width factors
            col_widths = []
            min_col_width = 0.4 * inch
            max_col_width = 4.0 * inch
            padding = 0.1 * inch  # Cell padding
            
            for i, col in enumerate(columns):
                # Calculate header width
                header_text = col['header']
                header_width = pdfmetrics.stringWidth(header_text, bold_font_name, 10) + padding
                
                # Calculate maximum data width
                max_data_width = 0
                for j, row in enumerate(table_data):
                    if j == 0:  # Skip header as we already calculated it
                        continue
                    if i < len(row):
                        cell_text = str(row[i])
                        # Use Paragraph for wrapped text, else normal text
                        if col.get('wrap_text', False):
                            # For wrapped text, use a reasonable fixed width
                            text_width = min(2.0 * inch, len(cell_text) * 0.07 * inch)
                        else:
                            text_width = pdfmetrics.stringWidth(cell_text, font_name, 8)
                        max_data_width = max(max_data_width, text_width)
                
                # Add padding to data width
                max_data_width += padding
                
                # Determine final column width
                col_width = max(header_width, max_data_width)
                col_width = max(min_col_width, min(col_width, max_col_width))
                col_widths.append(col_width)
            
            # Scale if total exceeds available width
            total_width = sum(col_widths)
            if total_width > available_width:
                # Try to reduce wider columns first
                scale_factor = available_width / total_width
                col_widths = [width * scale_factor for width in col_widths]
                
                # Ensure no column is too small after scaling
                for i in range(len(col_widths)):
                    if col_widths[i] < min_col_width:
                        # Redistribute space
                        needed = min_col_width - col_widths[i]
                        col_widths[i] = min_col_width
                        # Reduce other columns proportionally
                        other_cols = [j for j in range(len(col_widths)) if j != i and col_widths[j] > min_col_width]
                        if other_cols:
                            reduce_per_col = needed / len(other_cols)
                            for j in other_cols:
                                col_widths[j] = max(min_col_width, col_widths[j] - reduce_per_col)
            
            return col_widths
            
        except Exception as e:
            # Fallback to simple calculation
            print(f"Advanced width calculation failed, using fallback: {e}")
            return self._calculate_column_widths(columns, table_data, available_width)        
        
    def _calculate_column_widths(self, columns, table_data, available_width):
        """Calculate dynamic column widths based on content"""
        # Minimum and maximum column widths
        min_col_width = 0.5 * inch
        max_col_width = 3.0 * inch
        
        # Estimate character widths (approximate)
        char_width = 0.06 * inch  # Approximate width per character
        
        col_widths = []
        
        for i, col in enumerate(columns):
            # Get header width
            header_text = col['header']
            header_width = len(header_text) * char_width * 1.2  # Add 20% for bold
            
            # Get maximum data width for this column
            max_data_width = 0
            for row in table_data[1:]:  # Skip header row
                cell_text = str(row[i]) if i < len(row) else ""
                cell_width = len(cell_text) * char_width
                max_data_width = max(max_data_width, cell_width)
            
            # Use the maximum of header and data width
            content_width = max(header_width, max_data_width)
            
            # Apply min/max constraints
            col_width = max(min_col_width, min(content_width, max_col_width))
            col_widths.append(col_width)
        
        # If total width exceeds available width, scale down proportionally
        total_width = sum(col_widths)
        if total_width > available_width:
            scale_factor = available_width / total_width
            col_widths = [width * scale_factor for width in col_widths]
        
        return col_widths    
    

    def _auto_adjust_table_properties(self, columns, data, landscape_mode):
        """Automatically adjust table properties based on data"""
        num_columns = len(columns)
        total_data_rows = len(data)
        
        # Determine if we need landscape mode
        if num_columns > 6 and not landscape_mode:
            landscape_mode = True
            print("Auto-switching to landscape mode due to many columns")
        
        # Determine font size based on column count
        font_size = 10
        if num_columns > 8:
            font_size = 8
        elif num_columns > 12:
            font_size = 7
        
        return landscape_mode, font_size
    
    def _generate_pdf(self, data, columns, filename_prefix, start_date, end_date, 
                 title=None, row_styler=None, col_widths=None, landscape_mode=False):
        """Generate PDF report from data with company header"""
        try:
            # Auto-adjust table properties based on data
            landscape_mode, font_size = self._auto_adjust_table_properties(columns, data, landscape_mode)
            
            # Fetch company information from database
            company_info = self._get_company_info()
            
            # Create PDF in memory
            buffer = io.BytesIO()
            page_size = landscape(A4) if landscape_mode else A4
            
            # Calculate available width based on page size and margins
            if landscape_mode:
                available_width = page_size[0] - 1.0 * inch  # 0.5 inch margins on both sides
            else:
                available_width = page_size[0] - 1.0 * inch  # 0.5 inch margins on both sides
                
            doc = SimpleDocTemplate(buffer, pagesize=page_size, topMargin=0.5*inch, 
                                bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()

            # Create custom styles
            company_name_style = ParagraphStyle(
                'CompanyNameStyle',
                parent=styles['Normal'],
                textColor=colors.HexColor('#2B78E4'),
                fontSize=24,
                fontName='Helvetica-Bold',
                alignment=0,  # Left aligned
                spaceAfter=36
            )
            
            company_details_style = ParagraphStyle(
                'CompanyDetailsStyle',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica',
                alignment=0,  # Left aligned
                spaceAfter=2
            )
            
            report_title_style = ParagraphStyle(
                'ReportTitleStyle',
                parent=styles['Title'],
                fontSize=16,
                spaceBefore=12,
                spaceAfter=12,
                alignment=1  # Center aligned
            )

            # Create header table with logo and company info
            header_table_data = []
            
            # Logo cell (left side)
            logo_cell = []
            if company_info and company_info.get('logo_url'):
                try:
                    # Download and resize logo
                    response = requests.get(company_info['logo_url'], timeout=10)
                    if response.status_code == 200:
                        logo = Image(io.BytesIO(response.content), width=1.5*inch, height=0.5*inch)
                        logo.hAlign = 'LEFT'
                        logo_cell.append(logo)
                    else:
                        # Fallback if logo download fails
                        logo_cell.append(Paragraph("Company Logo", styles['Normal']))
                except Exception as logo_error:
                    print(f"Logo error: {logo_error}")
                    logo_cell.append(Paragraph("Company Logo", styles['Normal']))
            else:
                logo_cell.append(Paragraph("Company Logo", styles['Normal']))
            
            # Company details cell (right side)
            details_cell = []
            if company_info:
                # Company name
                if company_info.get('company_name'):
                    details_cell.append(Paragraph(company_info['company_name'], company_name_style))
                
                # GSTIN
                if company_info.get('gstin'):
                    details_cell.append(Paragraph(f"<b>GSTIN:</b> {company_info['gstin']}", company_details_style))
                
                # Phone
                if company_info.get('phone'):
                    details_cell.append(Paragraph(f"<b>Phone:</b> {company_info['phone']}", company_details_style))
                
                # Email
                if company_info.get('email'):
                    details_cell.append(Paragraph(f"<b>Email:</b> {company_info['email']}", company_details_style))
                
                # Address (with text wrapping)
                if company_info.get('address'):
                    address_style = ParagraphStyle(
                        'AddressStyle',
                        parent=company_details_style,
                        wordWrap='LTR',  # Enable text wrapping
                        maxLineLength=35  # Adjust based on your needs
                    )
                    details_cell.append(Paragraph(f"<b>Address:</b> {company_info['address']}", address_style))
            else:
                # Fallback if no company info
                details_cell.append(Paragraph("SMART MANUFACTURING ERP", company_name_style))

                details_cell.append(Paragraph("Company Information Not Available", company_details_style))
            
            # Create header table with two columns
            header_table_data = [
                [logo_cell, details_cell]
            ]
            
            header_table = Table(header_table_data, colWidths=[2*inch, 4*inch])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            elements.append(header_table)
            
            # Add separator line
            elements.append(Spacer(1, 0.1*inch))
            elements.append(self._create_line())
            elements.append(Spacer(1, 0.1*inch))
            
            # Add report title and date range
            report_title = title or f"{filename_prefix.replace('_', ' ').title()} Report"
            title_para = Paragraph(f"{report_title}", report_title_style)
            elements.append(title_para)
            
            # Date range
            date_range_style = ParagraphStyle(
                'DateRangeStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1,  # Center
                spaceAfter=12
            )
            date_para = Paragraph(f"Period: {start_date} to {end_date}", date_range_style)
            elements.append(date_para)
            
            # Add some space before the data table
            elements.append(Spacer(1, 0.2*inch))
            
            # Prepare table data
            table_data = [[col['header'] for col in columns]]
            
            # Track rows for styling
            styled_rows = []
            
            for i, row in enumerate(data, start=1):
                table_row = []
                for col in columns:
                    value = col['formatter'](row[col['key']]) if 'formatter' in col else str(row[col['key']] or '')
                    # Use Paragraph for text wrapping if specified
                    if col.get('wrap_text', False):
                        table_row.append(Paragraph(value, styles['Normal']))
                    else:
                        table_row.append(value)
                table_data.append(table_row)
                styled_rows.append(i)
            
            # Calculate dynamic column widths if not provided
            if not col_widths:
                col_widths = self._calculate_column_widths_advanced(columns, table_data, available_width, styles)
            
            # Create table
            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            # Base table style with dynamic font sizes
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2B78E4')),  # Company theme color
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), max(8, font_size)),  # Dynamic font size
                ('FONTSIZE', (0, 1), (-1, -1), max(7, font_size-1)),  # Smaller for data
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),  # Lighter grid
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]
            
            # Apply row styling if provided
            if row_styler:
                table_style = row_styler(table_style, data, styled_rows)
            
            table.setStyle(TableStyle(table_style))
            elements.append(table)
            
            # Add footer with generation timestamp
            elements.append(Spacer(1, 0.3*inch))
            footer_style = ParagraphStyle(
                'FooterStyle',
                parent=styles['Normal'],
                fontSize=8,
                alignment=2,  # Right aligned
                textColor=colors.grey
            )
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            footer_para = Paragraph(f"Generated on: {timestamp}", footer_style)
            elements.append(footer_para)
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{start_date}_to_{end_date}_{timestamp}.pdf"
            
            return Response(
                buffer,
                mimetype="application/pdf",
                headers={
                    "Content-Disposition": f"attachment;filename={filename}",
                    "Content-Type": "application/pdf"
                }
            )
        except Exception as e:
            print("Error generating PDF:", e)
            return {"errFlag": 1, "message": f"Error generating PDF: {str(e)}"}
        

    def _get_company_info(self):
        """Fetch company information from database"""
        try:
            sql = text('''
                SELECT 
                    company_name,
                    gstin,
                    phone,
                    email,
                    address,
                    logo_url
                FROM company_info_settings 
                WHERE status = 1 
                ORDER BY id DESC 
                LIMIT 1
            ''')
            
            with db.engine.connect() as conn:
                result = conn.execute(sql)
                company_data = result.mappings().first()
            
            if company_data:
                return dict(company_data)
            else:
                return None
                
        except Exception as e:
            print("Error fetching company info:", e)
            return None    
        
    def _create_line(self):
        """Create a horizontal line for separation"""
        from reportlab.lib.units import inch
        from reportlab.platypus import Flowable
        
        class Line(Flowable):
            def __init__(self, width, height=1):
                Flowable.__init__(self)
                self.width = width
                self.height = height

            def draw(self):
                self.canv.setStrokeColor(colors.HexColor('#2B78E4'))
                self.canv.setLineWidth(1)
                self.canv.line(0, 0, self.width, 0)
        
        return Line(7.5*inch)  # Standard width for A4

    # Common Formatters
    def _currency_formatter(self, value):
        """Format currency values"""
        try:
            return f"${float(value):.2f}"
        except (ValueError, TypeError):
            return "$0.00"
    
    def _date_formatter(self, value):
        """Format date values"""
        if isinstance(value, (datetime,)):
            return value.strftime('%Y-%m-%d')
        return str(value) if value else 'N/A'
    
    def _datetime_formatter(self, value):
        """Format datetime values"""
        if isinstance(value, (datetime,)):
            return value.strftime('%Y-%m-%d %H:%M')
        return str(value) if value else 'N/A'
    
    def _float_formatter(self, value, precision=2):
        """Format float values"""
        try:
            return f"{float(value):.{precision}f}"
        except (ValueError, TypeError):
            return f"0.{'0' * precision}"

    # Raw Materials Stock Report
    def generateRawMaterialsStockReport(self, start_date, end_date, doc_type='csv'):
        """Generates stock report for raw materials in specified format"""
        try:
            sql = text('''
                SELECT 
                    rm.material_code,
                    rm.material_name,
                    rm.material_description,
                    rm.stock_qty,
                    rm.min_stock_level,
                    rm.max_stock_level,
                    rm.unit_of_measure,
                    rm.unit_cost,
                    rm.total_value,
                    rm.stock_status,
                    rm.raw_material_image,
                    rm.last_restocked,
                    rm.created_at,
                    rm.updated_at
                FROM raw_materials rm
                WHERE (rm.created_at BETWEEN :start_date AND :end_date
                    OR rm.updated_at BETWEEN :start_date AND :end_date
                    OR rm.last_restocked BETWEEN :start_date AND :end_date)
                    AND rm.status = 1
                ORDER BY rm.material_name
            ''')
            
            with db.engine.connect() as conn:
                responseData = conn.execute(sql, {'start_date': start_date, 'end_date': end_date})
                results = responseData.mappings().all()
            
            if not results:
                return {"errFlag": 1, "message": "No data found for the selected date range"}
            
            # Process data
            report_data = []
            for row in results:
                stock_qty = row['stock_qty']
                min_stock = row['min_stock_level']
                low_stock_alert = "LOW STOCK" if stock_qty < min_stock else "OK"
                
                report_data.append({
                    'material_code': row['material_code'],
                    'material_name': row['material_name'],
                    'material_description': row['material_description'] or '',
                    'stock_qty': stock_qty,
                    'min_stock_level': min_stock,
                    'max_stock_level': row['max_stock_level'],
                    'unit_of_measure': row['unit_of_measure'],
                    'unit_cost': float(row['unit_cost']) if row['unit_cost'] else 0.0,
                    'total_value': float(row['total_value']) if row['total_value'] else 0.0,
                    'stock_status': row['stock_status'],
                    'low_stock_alert': low_stock_alert
                })
            
            # Define columns for report
            columns = [
                {'header': 'Material Code', 'key': 'material_code'},
                {'header': 'Material Name', 'key': 'material_name'},
                {'header': 'Description', 'key': 'material_description'},
                {'header': 'Stock Quantity', 'key': 'stock_qty'},
                {'header': 'Min Stock Level', 'key': 'min_stock_level'},
                {'header': 'Max Stock Level', 'key': 'max_stock_level'},
                {'header': 'Unit of Measure', 'key': 'unit_of_measure'},
                {'header': 'Unit Cost', 'key': 'unit_cost', 'formatter': self._currency_formatter},
                {'header': 'Total Value', 'key': 'total_value', 'formatter': self._currency_formatter},
                {'header': 'Stock Status', 'key': 'stock_status'},
                {'header': 'Low Stock Alert', 'key': 'low_stock_alert'}
            ]
            
            # Generate report based on document type
            if doc_type == 'csv':
                return self._generate_csv(report_data, columns, "raw_materials_stock", start_date, end_date)
            elif doc_type == 'excel':
                return self._generate_excel(report_data, columns, "raw_materials_stock", start_date, end_date)
            elif doc_type == 'pdf':
                def pdf_row_styler(table_style, data, styled_rows):
                    for i, row_idx in enumerate(styled_rows):
                        if data[i]['stock_qty'] < data[i]['min_stock_level']:
                            table_style.append(('TEXTCOLOR', (3, row_idx), (3, row_idx), colors.red))
                            table_style.append(('FONTNAME', (3, row_idx), (3, row_idx), 'Helvetica-Bold'))
                    return table_style
                
                return self._generate_pdf(
                    report_data, columns[:8],  # Show only first 8 columns in PDF for better fit
                    "raw_materials_stock", start_date, end_date,
                    title="Raw Materials Stock Report",
                    row_styler=pdf_row_styler
                )
            else:
                return {"errFlag": 1, "message": "Unsupported document type"}
            
        except Exception as e:
            print("Error generating stock report:", e)
            return {"errFlag": 1, "message": f"Error generating stock report: {str(e)}"}
    
    # Finished Goods Stock Report
    def generateFinishedGoodsStockReport(self, start_date, end_date, doc_type='csv'):
        """Generates stock report for finished goods in specified format"""
        try:
            sql = text('''
                SELECT
                    fg.id,
                    fg.product_name,
                    fg.sku_code,
                    fg.brand_id,
                    fg.product_category_id,
                    fg.stock_qty,
                    fg.min_level,
                    fg.max_level,
                    fg.unit_price,
                    fg.total_value,
                    fg.raw_material_cost,
                    fg.velocity,
                    fg.goods_status,
                    fg.last_produced,
                    fg.created_at,
                    fg.updated_at
                FROM finished_goods fg
                WHERE (fg.created_at BETWEEN :start_date AND :end_date
                    OR fg.updated_at BETWEEN :start_date AND :end_date
                    OR fg.last_produced BETWEEN :start_date AND :end_date)
                    AND fg.status = 1
                ORDER BY fg.product_name
            ''')

            with db.engine.connect() as conn:
                response = conn.execute(sql, {'start_date': start_date, 'end_date': end_date})
                results = response.mappings().all()

            if not results:
                return {"errFlag": 1, "message": "No finished goods found for the selected date range"}

            report_data = []
            for r in results:
                stock_qty = float(r['stock_qty'] or 0)
                min_level = float(r['min_level'] or 0)

                if stock_qty <= 0:
                    low_stock_alert = "OUT OF STOCK"
                elif min_level and stock_qty < min_level:
                    low_stock_alert = "LOW STOCK"
                else:
                    low_stock_alert = "OK"

                report_data.append({
                    'id': r['id'],
                    'product_name': r['product_name'],
                    'sku_code': r['sku_code'] or '',
                    'stock_qty': stock_qty,
                    'min_level': min_level,
                    'max_level': float(r['max_level'] or 0),
                    'unit_price': float(r['unit_price'] or 0.0),
                    'total_value': float(r['total_value'] or 0.0),
                    'raw_material_cost': float(r['raw_material_cost'] or 0.0),
                    'velocity': r['velocity'] or '',
                    'goods_status': r['goods_status'] or '',
                    'low_stock_alert': low_stock_alert,
                    'last_produced': r['last_produced'],
                })

            # Define columns for report
            columns = [
                {'header': 'ID', 'key': 'id'},
                {'header': 'Product Name', 'key': 'product_name'},
                {'header': 'SKU', 'key': 'sku_code'},
                {'header': 'Stock Quantity', 'key': 'stock_qty', 'formatter': lambda x: self._float_formatter(x, 2)},
                {'header': 'Min Level', 'key': 'min_level', 'formatter': lambda x: self._float_formatter(x, 2)},
                {'header': 'Max Level', 'key': 'max_level', 'formatter': lambda x: self._float_formatter(x, 2)},
                {'header': 'Unit Price', 'key': 'unit_price', 'formatter': self._currency_formatter},
                {'header': 'Total Value', 'key': 'total_value', 'formatter': self._currency_formatter},
                {'header': 'Raw Material Cost', 'key': 'raw_material_cost', 'formatter': self._currency_formatter},
                {'header': 'Velocity', 'key': 'velocity'},
                {'header': 'Goods Status', 'key': 'goods_status'},
                {'header': 'Low Stock Alert', 'key': 'low_stock_alert'},
                {'header': 'Last Produced', 'key': 'last_produced', 'formatter': self._date_formatter}
            ]

            if doc_type == 'csv':
                return self._generate_csv(report_data, columns, "finished_goods_stock", start_date, end_date)
            elif doc_type == 'excel':
                def excel_row_styler(worksheet, data, columns):
                    # Determine column indices
                    header = {cell.value: idx+1 for idx, cell in enumerate(worksheet[1])}
                    low_col = header.get('Low Stock Alert')
                    status_col = header.get('Goods Status')
                    
                    # Define fills
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    
                    for r_idx in range(2, worksheet.max_row + 1):
                        low_val = worksheet.cell(row=r_idx, column=low_col).value if low_col else None
                        status_val = worksheet.cell(row=r_idx, column=status_col).value if status_col else None
                        
                        if low_val == 'OUT OF STOCK' or (status_val and str(status_val).lower() == 'out-of-stock'):
                            fill = red_fill
                        elif low_val == 'LOW STOCK' or (status_val and str(status_val).lower() == 'low-stock'):
                            fill = yellow_fill
                        else:
                            fill = green_fill
                        
                        for c in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=r_idx, column=c).fill = fill
                
                return self._generate_excel(
                    report_data, columns, "finished_goods_stock", start_date, end_date,
                    row_styler=excel_row_styler
                )
            elif doc_type == 'pdf':
                def pdf_row_styler(table_style, data, styled_rows):
                    for i, row_idx in enumerate(styled_rows):
                        d = data[i]
                        if d['low_stock_alert'] == 'OUT OF STOCK' or (d['goods_status'] and d['goods_status'].lower() == 'out-of-stock'):
                            color = colors.HexColor("#FFC7CE")
                        elif d['low_stock_alert'] == 'LOW STOCK' or (d['goods_status'] and d['goods_status'].lower() == 'low-stock'):
                            color = colors.HexColor("#FFEB9C")
                        else:
                            color = colors.HexColor("#C6EFCE")
                        table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), color))
                    return table_style
                
                # Use only key columns for PDF
                pdf_columns = [
                    {'header': 'ID', 'key': 'id'},
                    {'header': 'Product Name', 'key': 'product_name', 'wrap_text': True},
                    {'header': 'SKU', 'key': 'sku_code'},
                    {'header': 'Stock Qty', 'key': 'stock_qty', 'formatter': lambda x: self._float_formatter(x, 2)},
                    {'header': 'Min Level', 'key': 'min_level', 'formatter': lambda x: self._float_formatter(x, 2)},
                    {'header': 'Unit Price', 'key': 'unit_price', 'formatter': self._currency_formatter},
                    {'header': 'Total Value', 'key': 'total_value', 'formatter': self._currency_formatter},
                    {'header': 'Goods Status', 'key': 'goods_status'}
                ]
                
                return self._generate_pdf(
                    report_data, pdf_columns, "finished_goods_stock", start_date, end_date,
                    title="Finished Goods Stock Report",
                    row_styler=pdf_row_styler
                )
            else:
                return {"errFlag": 1, "message": "Unsupported document type"}

        except Exception as e:
            print("Error generating finished goods report:", e)
            return {"errFlag": 1, "message": f"Error generating finished goods report: {str(e)}"}

    # Material Inward Report
    def generateMaterialInwardReport(self, start_date, end_date, doc_type='csv'):
        """Generates a detailed report of all items received from vendors"""
        try:
            sql = text('''
                SELECT
                    v.vendor_name,
                    vsr.received_date,
                    vsr.invoice_number,
                    rm.material_code,
                    rm.material_name,
                    vsi.received_qty,
                    vsi.unit_cost,
                    vsi.total_cost,
                    vsi.expiry_date
                FROM
                    vendor_stock_receipts AS vsr
                JOIN
                    vendors AS v ON vsr.vendor_id = v.id
                JOIN
                    vendor_stock_receipt_items AS vsi ON vsr.id = vsi.receipt_id
                JOIN
                    raw_materials AS rm ON vsi.raw_material_id = rm.id
                WHERE
                    vsr.received_date BETWEEN :start_date AND :end_date
                    AND vsr.status = 1
                    AND vsi.status = 1
                ORDER BY
                    v.vendor_name, vsr.received_date, rm.material_name
            ''')
            
            with db.engine.connect() as conn:
                results = conn.execute(sql, {'start_date': start_date, 'end_date': end_date}).mappings().all()

            if not results:
                return {"errFlag": 1, "message": "No vendor receipts found for the selected date range"}

            report_data = [dict(row) for row in results]

            # Define columns for report
            columns = [
                {'header': 'Vendor Name', 'key': 'vendor_name', 'wrap_text': True},
                {'header': 'Received Date', 'key': 'received_date', 'formatter': self._date_formatter},
                {'header': 'Invoice Number', 'key': 'invoice_number'},
                {'header': 'Material Code', 'key': 'material_code'},
                {'header': 'Material Name', 'key': 'material_name', 'wrap_text': True},
                {'header': 'Received Qty', 'key': 'received_qty', 'formatter': lambda x: self._float_formatter(x, 2)},
                {'header': 'Unit Cost', 'key': 'unit_cost', 'formatter': self._currency_formatter},
                {'header': 'Total Cost', 'key': 'total_cost', 'formatter': self._currency_formatter},
                {'header': 'Expiry Date', 'key': 'expiry_date', 'formatter': self._date_formatter}
            ]

            if doc_type == 'csv':
                return self._generate_csv(report_data, columns, "material_inward", start_date, end_date)
            elif doc_type == 'excel':
                return self._generate_excel(report_data, columns, "material_inward", start_date, end_date)
            elif doc_type == 'pdf':
                col_widths = [1.8*inch, 0.9*inch, 1*inch, 1*inch, 2.5*inch, 0.6*inch, 0.8*inch, 0.9*inch, 0.9*inch]
                return self._generate_pdf(
                    report_data, columns, "material_inward", start_date, end_date,
                    title="Material Inward Report",
                    col_widths=col_widths,
                    landscape_mode=True
                )
            else:
                return {"errFlag": 1, "message": "Unsupported document type"}

        except Exception as e:
            print("Error generating material inward report:", e)
            return {"errFlag": 1, "message": f"Error generating material inward report: {str(e)}"}

    # Quality Control Records Report
    def generateQcRecordsReport(self, start_date, end_date, doc_type='csv'):
        """Generates a report for Quality Control (QC) records"""
        try:
            sql = text('''
                SELECT
                    qcr.qc_code,
                    qcr.entity_type,
                    qcr.item_name,
                    qcr.inspector_name,
                    qct.test_type_name,
                    qcr.test_parameters,
                    qcr.remarks,
                    qcr.result,
                    qcr.status,
                    qcr.created_at
                FROM 
                    qc_records AS qcr
                LEFT JOIN 
                    qc_test_type AS qct ON qcr.test_type_id = qct.id
                WHERE 
                    qcr.created_at BETWEEN :start_date AND :end_date
                    AND qcr.status = 1
                ORDER BY 
                    qcr.created_at DESC
            ''')
            
            with db.engine.connect() as conn:
                results = conn.execute(sql, {'start_date': start_date, 'end_date': end_date}).mappings().all()

            if not results:
                return {"errFlag": 1, "message": "No QC records found for the selected date range"}

            report_data = [dict(row) for row in results]

            # Define columns for report
            columns = [
                {'header': 'QC Code', 'key': 'qc_code'},
                {'header': 'Entity Type', 'key': 'entity_type'},
                {'header': 'Item Name', 'key': 'item_name', 'wrap_text': True},
                {'header': 'Inspector Name', 'key': 'inspector_name'},
                {'header': 'Test Type', 'key': 'test_type_name'},
                {'header': 'Test Parameters', 'key': 'test_parameters', 'wrap_text': True},
                {'header': 'Result', 'key': 'result'},
                {'header': 'Remarks', 'key': 'remarks', 'wrap_text': True},
                {'header': 'Date', 'key': 'created_at', 'formatter': self._datetime_formatter}
            ]

            if doc_type == 'csv':
                return self._generate_csv(report_data, columns, "qc_records", start_date, end_date)
            elif doc_type == 'excel':
                return self._generate_excel(report_data, columns, "qc_records", start_date, end_date)
            elif doc_type == 'pdf':
                col_widths = [1.0*inch, 0.8*inch, 2.0*inch, 1.1*inch, 1.4*inch, 1.5*inch, 0.7*inch, 2.5*inch, 1.1*inch]
                return self._generate_pdf(
                    report_data, columns, "qc_records", start_date, end_date,
                    title="Quality Control Records Report",
                    col_widths=col_widths,
                    landscape_mode=True
                )
            else:
                return {"errFlag": 1, "message": "Unsupported document type"}

        except Exception as e:
            print("Error generating QC records report:", e)
            return {"errFlag": 1, "message": f"Error generating QC records report: {str(e)}"}

    # Raw Material Consumption Report
    def generateRawMaterialConsumptionReport(self, start_date, end_date, doc_type='csv'):
        """Generates raw material consumption report for given date range"""
        try:
            sql = text('''
                SELECT
                    r.id,
                    r.production_batch_id,
                    pb.production_code AS production_code,
                    r.raw_material_id,
                    rm.material_code,
                    rm.material_name,
                    r.consumed_qty,
                    r.unit,
                    r.unit_cost,
                    r.total_cost,
                    r.notes,
                    r.created_at,
                    r.created_admin_id,
                    r.status
                FROM raw_material_consumption_receipt r
                LEFT JOIN raw_materials rm ON rm.id = r.raw_material_id
                LEFT JOIN production_batch pb ON pb.id = r.production_batch_id
                WHERE r.created_at BETWEEN :start_date AND :end_date
                ORDER BY r.created_at ASC
            ''')

            with db.engine.connect() as conn:
                resp = conn.execute(sql, {'start_date': start_date, 'end_date': end_date})
                rows = resp.mappings().all()

            if not rows:
                return {"errFlag": 1, "message": "No consumption records found for the selected date range"}

            report_data = []
            for r in rows:
                report_data.append({
                    'id': r['id'],
                    'production_batch_id': r['production_batch_id'],
                    'production_code': r.get('production_code') or '',
                    'raw_material_id': r['raw_material_id'],
                    'material_code': r.get('material_code') or '',
                    'material_name': r.get('material_name') or '',
                    'consumed_qty': float(r['consumed_qty']) if r['consumed_qty'] is not None else 0.0,
                    'unit': r.get('unit') or '',
                    'unit_cost': float(r['unit_cost']) if r['unit_cost'] is not None else 0.0,
                    'total_cost': float(r['total_cost']) if r['total_cost'] is not None else 0.0,
                    'notes': r.get('notes') or '',
                    'created_at': r['created_at'],
                    'created_admin_id': r.get('created_admin_id'),
                    'status': r.get('status', 1)
                })

            # Define columns for report
            columns = [
                {'header': 'ID', 'key': 'id'},
                {'header': 'Production Batch ID', 'key': 'production_batch_id'},
                {'header': 'Production Code', 'key': 'production_code'},
                {'header': 'Material Code', 'key': 'material_code'},
                {'header': 'Material Name', 'key': 'material_name', 'wrap_text': True},
                {'header': 'Consumed Qty', 'key': 'consumed_qty', 'formatter': lambda x: self._float_formatter(x, 4)},
                {'header': 'Unit', 'key': 'unit'},
                {'header': 'Unit Cost', 'key': 'unit_cost', 'formatter': self._currency_formatter},
                {'header': 'Total Cost', 'key': 'total_cost', 'formatter': self._currency_formatter},
                {'header': 'Notes', 'key': 'notes', 'wrap_text': True},
                {'header': 'Date', 'key': 'created_at', 'formatter': self._datetime_formatter}
            ]

            if doc_type == 'csv':
                return self._generate_csv(report_data, columns, "raw_material_consumption", start_date, end_date)
            elif doc_type == 'excel':
                return self._generate_excel(report_data, columns, "raw_material_consumption", start_date, end_date)
            elif doc_type == 'pdf':
                def pdf_row_styler(table_style, data, styled_rows):
                    for i, row_idx in enumerate(styled_rows):
                        if data[i]['consumed_qty'] < 0:
                            table_style.append(('TEXTCOLOR', (5, row_idx), (5, row_idx), colors.red))
                            table_style.append(('FONTNAME', (5, row_idx), (5, row_idx), 'Helvetica-Bold'))
                    return table_style
                
                # Use only key columns for PDF
                pdf_columns = [
                    {'header': 'ID', 'key': 'id'},
                    {'header': 'Batch', 'key': 'production_batch_id'},
                    {'header': 'Prod Code', 'key': 'production_code'},
                    {'header': 'Material', 'key': 'material_code'},
                    {'header': 'Material Name', 'key': 'material_name', 'wrap_text': True},
                    {'header': 'Qty', 'key': 'consumed_qty', 'formatter': lambda x: self._float_formatter(x, 4)},
                    {'header': 'Unit', 'key': 'unit'},
                    {'header': 'Unit Cost', 'key': 'unit_cost', 'formatter': self._currency_formatter},
                    {'header': 'Total Cost', 'key': 'total_cost', 'formatter': self._currency_formatter},
                    {'header': 'Date', 'key': 'created_at', 'formatter': self._datetime_formatter}
                ]
                
                return self._generate_pdf(
                    report_data, pdf_columns, "raw_material_consumption", start_date, end_date,
                    title="Raw Material Consumption Report",
                    row_styler=pdf_row_styler
                )
            else:
                return {"errFlag": 1, "message": "Unsupported document type"}

        except Exception as e:
            print("Error generating consumption report:", e)
            return {"errFlag": 1, "message": f"Error generating consumption report: {str(e)}"}

    # Vendor Performance Report
    def generateVendorPerformanceReport(self, doc_type='csv'):
        """Generates a vendor performance report"""
        try:
            sql = text("""
                SELECT
                    v.id AS vendor_id, v.vendor_name, v.credit_limit, v.on_time_percentage,
                    rm.material_name
                FROM vendors v
                LEFT JOIN vendor_raw_materials vrm ON v.id = vrm.vendor_id
                LEFT JOIN raw_materials rm ON vrm.raw_material_id = rm.id
                WHERE v.status = 1
                ORDER BY v.vendor_name, rm.material_name;
            """)
            with db.engine.connect() as conn:
                results = conn.execute(sql).mappings().all()

            if not results:
                return {"errFlag": 1, "message": "No active vendors found to generate a report."}

            # Process the flat data into a structured list
            vendors_dict = {}
            for row in results:
                vendor_id = row['vendor_id']
                if vendor_id not in vendors_dict:
                    vendors_dict[vendor_id] = {
                        "vendor_name": row['vendor_name'],
                        "credit_limit": float(row['credit_limit'] or 0.0),
                        "on_time_percentage": float(row['on_time_percentage'] or 0.0),
                        "materials_supplied": []
                    }
                if row['material_name']:
                    vendors_dict[vendor_id]['materials_supplied'].append(row['material_name'])
            
            report_data = list(vendors_dict.values())
            current_date = datetime.now().strftime("%Y-%m-%d")

            # Define columns for report
            columns = [
                {'header': 'Vendor Name', 'key': 'vendor_name', 'wrap_text': True},
                {'header': 'Credit Limit', 'key': 'credit_limit', 'formatter': self._currency_formatter},
                {'header': 'On-Time Delivery (%)', 'key': 'on_time_percentage', 'formatter': lambda x: self._float_formatter(x, 2)},
                {'header': 'Materials Supplied', 'key': 'materials_supplied', 'formatter': lambda x: ", ".join(x) if x else "N/A", 'wrap_text': True}
            ]

            if doc_type == 'csv':
                return self._generate_csv(report_data, columns, "vendor_performance", current_date, current_date)
            elif doc_type == 'excel':
                return self._generate_excel(report_data, columns, "vendor_performance", current_date, current_date)
            elif doc_type == 'pdf':
                col_widths = [2.5*inch, 1.5*inch, 1.5*inch, 5*inch]
                return self._generate_pdf(
                    report_data, columns, "vendor_performance", current_date, current_date,
                    title="Vendor Performance Report",
                    col_widths=col_widths,
                    landscape_mode=True
                )
            else:
                return {"errFlag": 1, "message": "Unsupported document type"}

        except Exception as e:
            print(f"Error generating vendor performance report: {e}")
            return {"errFlag": 1, "message": f"Error generating report: {str(e)}"}

# Create a single instance of the class to be used by the routes
reportsObj = ReportsClass()